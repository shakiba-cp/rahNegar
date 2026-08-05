# -*- coding: utf-8 -*-
"""
ره‌نگار — سامانه مدیریت فعالیت‌های امنیتی
نرم‌افزار تحت وب برای ثبت، مدیریت و گزارش‌گیری فعالیت‌های تیم امنیت.
اجرا:  python app.py   سپس  http://127.0.0.1:5000
"""
import csv
import datetime
import io
import json
import os
import re
import secrets
import sqlite3
from functools import wraps

from flask import (Flask, abort, flash, g, jsonify, redirect, render_template,
                   request, send_file, session, url_for)
from werkzeug.security import check_password_hash, generate_password_hash

import jalali

try:
    from openpyxl import Workbook, load_workbook
except ImportError:  # pragma: no cover
    Workbook = None
    load_workbook = None

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DB_PATH = os.path.join(BASE_DIR, "secman.db")
UPLOAD_DIR = os.path.join(BASE_DIR, "uploads")

STATUSES = ["در حال انجام", "انجام شده", "بررسی شده"]
EXPERT_STATUSES = ["در حال انجام", "انجام شده"]
DEFAULT_FORMATS = "pdf,doc,docx,xls,xlsx,csv,jpg,jpeg,png,zip"

# نگاشت حوزه‌ها به شناسه آیکن SVG (symbol id در اسپرایت base.html)
DOMAIN_ICONS = {
    "استخراج IP آسیب پذیر": "d-target", "رصد": "i-eye", "ارزیابی امنیتی وب": "i-globe",
    "اخبار پورتال ماهر": "d-news", "ارزیابی امنیتی اندروید": "d-phone", "بدافزار": "i-bug",
    "ابزار": "i-wrench", "دوره آموزشی": "i-cap", "مستندات فنی": "i-doc",
    "پایش و غربالگری": "d-scan", "خدمات مشاوره ای": "d-headset", "خدمات موردی": "d-clip",
    "شیفت آنکال": "d-call",
}

# (label, type, key, options, required, section) — section='' یعنی فیلدهای اصلی
SEC_REQUEST = "مشخصات درخواست‌دهنده فعالیت"
SEC_DELIVERY = "مشخصات تحویل فعالیت"

HOW = ["تلفنی", "ایمیل", "تیکت / سامانه", "حضوری", "مکاتبه"]
SEV = ["کم", "متوسط", "زیاد", "بحرانی"]
STATUSES_FIELD = ["در حال انجام", "انجام شده", "بررسی شده"]
VULN_ST = ["باز", "در حال رفع", "رفع‌شده", "ریسک پذیرفته‌شده"]

# بخش‌های مشترک همه حوزه‌ها (بعد از فیلدهای اصلی هر حوزه اضافه می‌شود)
COMMON_FIELDS = [
    ("زمان درخواست", "date", "request_date", [], 0, SEC_REQUEST),
    ("نام درخواست‌دهنده", "text", None, [], 0, SEC_REQUEST),
    ("نحوه درخواست فعالیت", "select", None, HOW, 0, SEC_REQUEST),
    ("مستندات مربوط به درخواست", "text", None, [], 0, SEC_REQUEST),
    ("زمان تحویل فعالیت", "date", "delivery_date", [], 0, SEC_DELIVERY),
    ("نحوه تحویل", "select", None, HOW, 0, SEC_DELIVERY),
    ("نام تحویل‌گیرنده", "text", None, [], 0, SEC_DELIVERY),
    ("مستندات تحویل‌شده", "text", None, [], 0, SEC_DELIVERY),
    ("نام تحویل‌دهنده", "text", None, [], 0, SEC_DELIVERY),
]

DOMAIN_FIELDS = {
    "استخراج IP آسیب پذیر": [
        ("تاریخ", "date", "date", [], 1, ""),
        ("کارشناس", "text", "expert", [], 1, ""),
        ("CVE", "text", None, [], 0, ""),
        ("محصول", "text", "title", [], 1, ""),
        ("حوزه", "select", None, ["دیتاسنتر", "حاکمیتی غیرزیرساختی", "پایه"], 1, ""),
        ("شدت", "select", None, SEV, 1, ""),
        ("توضیحات", "textarea", None, [], 0, ""),
    ],
    "رصد": [
        ("عنوان", "text", "title", [], 1, ""),
        ("وضعیت", "select", None, ["بررسی نشده", "در حال بررسی", "اطلاع‌رسانی شده", "بسته"], 0, ""),
        ("تاریخ", "date", "date", [], 1, ""),
        ("توضیحات", "textarea", None, [], 0, ""),
    ],
    "ارزیابی امنیتی وب": [
        ("شناسه", "text", None, [], 0, ""),
        ("شناسه فرآیند", "text", None, [], 0, ""),
        ("تاریخ", "date", "date", [], 1, ""),
        ("کارشناس", "text", "expert", [], 1, ""),
        ("کارفرما", "text", None, [], 1, ""),
        ("پروفایل", "select", None, ["پروفایل صفر", "پروفایل ۱", "پروفایل ۲", "پروفایل ۳",
                                      "پروفایل ۴", "پروفایل ۵", "پروفایل ۶"], 0, ""),
        ("آسیب پذیری", "textarea", None, [], 0, ""),
        ("شدت", "select", None, SEV, 1, ""),
        ("وضعیت آسیب‌پذیری", "select", None, ["تایید", "عدم تایید"], 0, ""),
        ("آدرس", "text", "title", [], 0, ""),
        ("توضیحات", "textarea", None, [], 0, ""),
    ],
    "اخبار پورتال ماهر": [
        ("CVE", "text", "title", [], 1, ""),
        ("کارشناس", "text", "expert", [], 1, ""),
        ("محصول", "text", None, [], 0, ""),
        ("تاریخ", "date", "date", [], 1, ""),
        ("توضیحات", "textarea", None, [], 0, ""),
    ],
    "ارزیابی امنیتی اندروید": [
        ("شناسه", "text", None, [], 0, ""),
        ("تاریخ", "date", "date", [], 1, ""),
        ("کارشناس", "text", "expert", [], 1, ""),
        ("کارفرما", "text", "title", [], 1, ""),
        ("برنامه", "text", None, [], 1, ""),
        ("آسیب پذیری", "textarea", None, [], 0, ""),
        ("شدت", "select", None, SEV, 1, ""),
        ("وضعیت آسیب‌پذیری", "select", None, VULN_ST, 0, ""),
        ("توضیحات", "textarea", None, [], 0, ""),
    ],
    "بدافزار": [
        ("بدافزار", "text", "title", [], 1, ""),
        ("کارشناس", "text", "expert", [], 1, ""),
        ("تاریخ", "date", "date", [], 1, ""),
        ("توضیحات", "textarea", None, [], 0, ""),
    ],
    "ابزار": [
        ("نام ابزار", "text", "title", [], 1, ""),
        ("تاریخ ارسال به مرکز ماهر", "date", "date", [], 1, ""),
        ("تحویل گیرندگان در مرکز ماهر", "text", None, [], 0, ""),
        ("شرح عملکرد ابزار", "textarea", None, [], 0, ""),
        ("مستندات فنی", "file", None, [], 0, ""),
        ("آدرس URL دسترسی", "text", None, [], 0, ""),
        ("تاریخ جلسات برگزار شده", "text", None, [], 0, ""),
        ("شرکت‌کنندگان جلسه", "textarea", None, [], 0, ""),
        ("مدت زمان جلسه", "text", None, [], 0, ""),
        ("توضیحات", "textarea", None, [], 0, ""),
    ],
    "دوره آموزشی": [
        ("نام دوره آموزشی", "text", "title", [], 1, ""),
        ("نحوه برگزاری دوره", "select", None, ["حضوری", "مجازی", "ترکیبی"], 0, ""),
        ("زمان شروع", "date", "date", [], 1, ""),
        ("زمان پایان", "date", None, [], 0, ""),
        ("مدرسان", "text", None, [], 0, ""),
        ("تعداد ساعت برگزاری", "number", None, [], 0, ""),
        ("درخواست دهنده", "text", None, [], 0, ""),
        ("تعداد شرکت‌کنندگان", "number", None, [], 0, ""),
        ("پوستر دوره", "file", None, [], 0, ""),
        ("لیست نفرات", "file", None, [], 0, ""),
        ("مستندات اضافی دوره", "file", None, [], 0, ""),
        ("توضیحات", "textarea", None, [], 0, ""),
    ],
    "مستندات فنی": [
        ("عنوان مستند", "text", "title", [], 1, ""),
        ("نوع مستند", "select", None, ["پروپزال", "گزارش فنی پروژه", "گزارش دوره‌ای", "سایر"], 1, ""),
        ("تاریخ ارائه مستند", "date", "date", [], 1, ""),
        ("ارائه کننده", "text", "expert", [], 1, ""),
        ("تحویل گیرنده در مرکز ماهر", "text", None, [], 0, ""),
        ("نحوه ارائه مستند به ماهر", "text", None, [], 0, ""),
        ("درخواست دهنده مستند", "text", None, [], 0, ""),
        ("توضیحات", "textarea", None, [], 0, ""),
    ],
    "پایش و غربالگری": [
        ("دستگاه ارزیابی‌شونده", "text", "title", [], 1, ""),
        ("دستگاه ارزیابی‌کننده", "text", None, [], 0, ""),
        ("طبقه‌بندی دستگاه", "text", None, [], 0, ""),
        ("تاریخ و ساعت مراجعه", "date", "date", [], 1, ""),
        ("شماره مکاتبه هماهنگی", "text", None, [], 0, ""),
        ("مسئول ارزیابی‌کننده", "text", "expert", [], 1, ""),
        ("عدم انجام کامل ارزیابی به دلیل", "select", None,
         ["انجام کامل شد", "عدم حضور مسئول فناوری اطلاعات", "ممانعت از ورود ارزیاب",
          "فاقد شبکه فناوری اطلاعات", "سایر موارد"], 0, ""),
        ("توضیحات (سایر موارد)", "textarea", None, [], 0, ""),
        ("مستندات فنی", "file", None, [], 0, ""),
    ],
    "خدمات مشاوره ای": [
        ("اطلاعات درخواست‌دهنده خدمات", "text", None, [], 0, ""),
        ("نام خدمات مشاوره‌ای", "text", "title", [], 1, ""),
        ("تاریخ ارائه خدمات", "date", "date", [], 1, ""),
        ("محل ارائه خدمات", "text", None, [], 0, ""),
        ("نام سازمان گیرنده خدمات", "text", None, [], 0, ""),
        ("نام شخص گیرنده خدمات", "text", None, [], 0, ""),
        ("نام ارائه‌کنندگان خدمات", "text", "expert", [], 0, ""),
        ("نحوه ارائه خدمات", "select", None, ["حضوری", "مجازی", "ترکیبی"], 0, ""),
        ("ساعت مفید کاری ارائه خدمات", "number", None, [], 0, ""),
        ("مستندات ارائه خدمات", "file", None, [], 0, ""),
        ("توضیحات", "textarea", None, [], 0, ""),
    ],
    "خدمات موردی": [
        ("درخواست‌کننده", "text", None, [], 1, ""),
        ("تاریخ", "date", "date", [], 1, ""),
        ("خدمت ارائه شده", "text", "title", [], 1, ""),
        ("شرح خدمت", "textarea", None, [], 0, ""),
        ("مستندات فنی", "file", None, [], 0, ""),
    ],
    "شیفت آنکال": [
        ("تاریخ درخواست", "date", None, [], 0, ""),
        ("درخواست‌دهنده", "text", "title", [], 1, ""),
        ("کانال درخواست", "select", None, HOW, 0, ""),
        ("نحوه شیفت", "select", None, ["حضوری", "مجازی", "ترکیبی"], 0, ""),
        ("تاریخ شروع", "date", "date", [], 1, ""),
        ("تاریخ پایان", "date", None, [], 0, ""),
        ("مستندات نهایی", "file", None, [], 0, ""),
        ("توضیحات", "textarea", None, [], 0, ""),
    ],
}

app = Flask(__name__)
# نسخه دارایی‌های استاتیک برای شکستن کش مرورگر بعد از هر آپدیت (cache-busting)
ASSET_V = "6.3.0"
app.config["SEND_FILE_MAX_AGE_DEFAULT"] = 0
# تقویت امنیت کوکی نشست — Secure را هنگام HTTPS با متغیر محیطی SECURE_COOKIE=1 فعال کنید
app.config["SESSION_COOKIE_HTTPONLY"] = True
app.config["SESSION_COOKIE_SAMESITE"] = "Lax"
# انقضای نشست: ۸ ساعت عدم فعالیت → خروج خودکار (با هر درخواست تمدید می‌شود)
app.config["PERMANENT_SESSION_LIFETIME"] = datetime.timedelta(hours=8)
if os.environ.get("SECURE_COOKIE") == "1":
    app.config["SESSION_COOKIE_SECURE"] = True


@app.after_request
def _security_headers(resp):
    """هدرهای تقویتی امنیت وب (clickjacking/sniffing/لو رفتن referrer)"""
    h = resp.headers
    h.setdefault("X-Frame-Options", "DENY")
    h.setdefault("X-Content-Type-Options", "nosniff")
    h.setdefault("Referrer-Policy", "same-origin")
    h.setdefault("Permissions-Policy", "camera=(), microphone=(), geolocation=()")
    h.setdefault("Content-Security-Policy",
                 "default-src 'self'; img-src 'self' data:; "
                 "style-src 'self' 'unsafe-inline'; script-src 'self' 'unsafe-inline'; "
                 "font-src 'self'; connect-src 'self'; form-action 'self'; "
                 "frame-ancestors 'none'; base-uri 'self'")
    return resp
# قالب‌های جینجا هنگام تغییر فایل، خودکار بازخوانی شوند (بدون نیاز به ری‌استارت)
app.config["TEMPLATES_AUTO_RELOAD"] = True


# ---------------------------------------------------------------- پایگاه داده
def get_db():
    if "db" not in g:
        g.db = sqlite3.connect(DB_PATH)
        g.db.row_factory = sqlite3.Row
        g.db.execute("PRAGMA foreign_keys = ON")
    return g.db


@app.teardown_appcontext
def _close_db(_exc):
    db = g.pop("db", None)
    if db is not None:
        db.close()


SCHEMA = """
CREATE TABLE IF NOT EXISTS users(
  id INTEGER PRIMARY KEY AUTOINCREMENT,
  username TEXT UNIQUE NOT NULL,
  password_hash TEXT NOT NULL,
  full_name TEXT NOT NULL,
  role TEXT NOT NULL DEFAULT 'expert',
  is_active INTEGER NOT NULL DEFAULT 1,
  aliases TEXT NOT NULL DEFAULT '',
  created_at TEXT
);
CREATE TABLE IF NOT EXISTS domains(
  id INTEGER PRIMARY KEY AUTOINCREMENT,
  name TEXT UNIQUE NOT NULL,
  sort_order INTEGER NOT NULL DEFAULT 0,
  is_active INTEGER NOT NULL DEFAULT 1,
  org_id INTEGER
);
CREATE TABLE IF NOT EXISTS orgs(
  id INTEGER PRIMARY KEY AUTOINCREMENT,
  name TEXT UNIQUE NOT NULL,
  sort_order INTEGER NOT NULL DEFAULT 0,
  is_active INTEGER NOT NULL DEFAULT 1
);
CREATE TABLE IF NOT EXISTS form_fields(
  id INTEGER PRIMARY KEY AUTOINCREMENT,
  domain_id INTEGER NOT NULL REFERENCES domains(id) ON DELETE CASCADE,
  label TEXT NOT NULL,
  field_key TEXT,
  field_type TEXT NOT NULL DEFAULT 'text',
  section TEXT NOT NULL DEFAULT '',
  options TEXT,
  required INTEGER NOT NULL DEFAULT 0,
  sort_order INTEGER NOT NULL DEFAULT 0,
  is_active INTEGER NOT NULL DEFAULT 1
);
CREATE TABLE IF NOT EXISTS activities(
  id INTEGER PRIMARY KEY AUTOINCREMENT,
  domain_id INTEGER NOT NULL REFERENCES domains(id),
  user_id INTEGER NOT NULL REFERENCES users(id),
  status TEXT NOT NULL DEFAULT 'در حال انجام',
  title TEXT,
  ticket TEXT,
  date TEXT,
  created_at TEXT,
  updated_at TEXT,
  created_by INTEGER,
  flagged INTEGER NOT NULL DEFAULT 0
);
CREATE TABLE IF NOT EXISTS responses(
  id INTEGER PRIMARY KEY AUTOINCREMENT,
  activity_id INTEGER NOT NULL REFERENCES activities(id) ON DELETE CASCADE,
  user_id INTEGER NOT NULL,
  body TEXT NOT NULL,
  created_at TEXT
);
CREATE TABLE IF NOT EXISTS activity_values(
  id INTEGER PRIMARY KEY AUTOINCREMENT,
  activity_id INTEGER NOT NULL REFERENCES activities(id) ON DELETE CASCADE,
  field_id INTEGER NOT NULL REFERENCES form_fields(id) ON DELETE CASCADE,
  value TEXT,
  UNIQUE(activity_id, field_id)
);
CREATE TABLE IF NOT EXISTS attachments(
  id INTEGER PRIMARY KEY AUTOINCREMENT,
  activity_id INTEGER NOT NULL REFERENCES activities(id) ON DELETE CASCADE,
  stored_name TEXT NOT NULL,
  original_name TEXT NOT NULL,
  size INTEGER DEFAULT 0,
  uploaded_by INTEGER,
  uploaded_at TEXT
);
CREATE TABLE IF NOT EXISTS excel_imports(
  id INTEGER PRIMARY KEY AUTOINCREMENT,
  domain_id INTEGER REFERENCES domains(id),
  user_id INTEGER REFERENCES users(id),
  filename TEXT,
  total_rows INTEGER, success_rows INTEGER, error_rows INTEGER,
  dup_rows INTEGER NOT NULL DEFAULT 0,
  warns TEXT NOT NULL DEFAULT '', errors TEXT, imported_at TEXT
);
CREATE TABLE IF NOT EXISTS settings(key TEXT PRIMARY KEY, value TEXT);
"""


def now_iso():
    return datetime.datetime.now().replace(microsecond=0).isoformat(" ")


def get_setting(key, default=""):
    row = get_db().execute("SELECT value FROM settings WHERE key=?", (key,)).fetchone()
    return row["value"] if row else default


def set_setting(key, value):
    db = get_db()
    db.execute("INSERT INTO settings(key,value) VALUES(?,?) "
               "ON CONFLICT(key) DO UPDATE SET value=excluded.value", (key, value))
    db.commit()


def allowed_formats():
    raw = get_setting("allowed_formats", DEFAULT_FORMATS)
    return [x.strip().lower().lstrip(".") for x in re.split(r"[,،;\s]+", raw)
            if x.strip()]


def max_upload_mb():
    try:
        return max(1, int(get_setting("max_upload_mb", "10")))
    except ValueError:
        return 10


def init_db():
    os.makedirs(UPLOAD_DIR, exist_ok=True)
    db = get_db()
    db.executescript(SCHEMA)
    if not get_setting("secret_key"):
        set_setting("secret_key", secrets.token_hex(32))
    app.secret_key = get_setting("secret_key")
    for k, v in {"system_name": "ره‌نگار",
                 "max_upload_mb": "10",
                 "allowed_formats": DEFAULT_FORMATS}.items():
        if get_setting(k) is None or get_setting(k) == "":
            set_setting(k, v)
    # مراکز (هر مرکز حوزه‌های مستقل خود را دارد) — افزودنی و سازگار با دیتابیس‌های قدیمی
    for _i, _nm in enumerate(["ماهر", "کاشف", "فیدار"], start=1):
        if not db.execute("SELECT 1 FROM orgs WHERE name=?", (_nm,)).fetchone():
            db.execute("INSERT INTO orgs(name,sort_order) VALUES(?,?)", (_nm, _i))
    if "org_id" not in [r[1] for r in db.execute("PRAGMA table_info(domains)")]:
        db.execute("ALTER TABLE domains ADD COLUMN org_id INTEGER")
    _mah = db.execute("SELECT id FROM orgs WHERE name='ماهر'").fetchone()["id"]
    db.execute("UPDATE domains SET org_id=? WHERE org_id IS NULL", (_mah,))
    # نام‌های مستعار کارشناسان (برای یکدست‌سازی نام در نمودارها/گزارش‌ها)
    if "aliases" not in [r[1] for r in db.execute("PRAGMA table_info(users)")]:
        db.execute("ALTER TABLE users ADD COLUMN aliases TEXT NOT NULL DEFAULT ''")
    if not db.execute("SELECT 1 FROM users WHERE username='admin'").fetchone():
        db.execute("INSERT INTO users(username,password_hash,full_name,role,created_at) "
                   "VALUES(?,?,?,?,?)",
                   ("admin", generate_password_hash("admin123"), "مدیر سامانه",
                    "admin", now_iso()))
    if not db.execute("SELECT 1 FROM domains LIMIT 1").fetchone():
        for i, name in enumerate(DOMAIN_FIELDS, start=1):
            cur = db.execute("INSERT INTO domains(name,sort_order,org_id) VALUES(?,?,?)",
                             (name, i, _mah))
            main_fields = [list(f) for f in DOMAIN_FIELDS[name]]
            if not any(f[2] == "ticket" for f in main_fields):
                main_fields.append(("شماره تیکت", "text", "ticket", [], 0, ""))
            for j, (label, ftype, fkey, opts, req, sec) in enumerate(
                    main_fields + COMMON_FIELDS, start=1):
                db.execute("INSERT INTO form_fields(domain_id,label,field_key,field_type,"
                           "section,options,required,sort_order) VALUES(?,?,?,?,?,?,?,?)",
                           (cur.lastrowid, label, fkey, ftype, sec,
                            json.dumps(opts, ensure_ascii=False), req, j))
    # مهاجرت نرم برای پایگاه‌های قدیمی: افزودن ستون section در صورت نبود
    try:
        db.execute("ALTER TABLE form_fields ADD COLUMN section TEXT NOT NULL DEFAULT ''")
    except sqlite3.OperationalError:
        pass
    try:
        db.execute("ALTER TABLE activities ADD COLUMN created_by INTEGER")
    except sqlite3.OperationalError:
        pass
    try:
        db.execute("ALTER TABLE activities ADD COLUMN task_note TEXT NOT NULL DEFAULT ''")
    except sqlite3.OperationalError:
        pass
    try:
        db.execute("ALTER TABLE activities ADD COLUMN flagged INTEGER NOT NULL DEFAULT 0")
    except sqlite3.OperationalError:
        pass
    # مهاجرت نرم: ردیف‌های تکراریِ نادیده‌گرفته‌شده در ورود Excel
    try:
        db.execute("ALTER TABLE excel_imports ADD COLUMN dup_rows INTEGER NOT NULL DEFAULT 0")
    except sqlite3.OperationalError:
        pass
    try:
        db.execute("ALTER TABLE excel_imports ADD COLUMN warns TEXT NOT NULL DEFAULT ''")
    except sqlite3.OperationalError:
        pass
    # مهاجرت‌های نرم v5.7: مجوزهای دانه‌ای + کارآموز/سرپرست
    for col, ddl in (("can_add", "INTEGER NOT NULL DEFAULT 1"),
                     ("can_edit", "INTEGER NOT NULL DEFAULT 1"),
                     ("can_delete", "INTEGER NOT NULL DEFAULT 0"),
                     ("can_import", "INTEGER NOT NULL DEFAULT 1"),
                     ("is_trainee", "INTEGER NOT NULL DEFAULT 0"),
                     ("supervisor_id", "INTEGER")):
        try:
            db.execute(f"ALTER TABLE users ADD COLUMN {col} {ddl}")
        except sqlite3.OperationalError:
            pass
    # گزینه‌های «وضعیت آسیب‌پذیری» حوزه ارزیابی امنیتی وب → تایید / عدم تایید
    try:
        db.execute("""UPDATE form_fields SET options='["تایید", "عدم تایید"]'
                      WHERE label='وضعیت آسیب‌پذیری' AND field_type='select'
                        AND domain_id=(SELECT id FROM domains
                                       WHERE name='ارزیابی امنیتی وب' LIMIT 1)""")
    except sqlite3.OperationalError:
        pass
    # افزودن فرمت csv به فهرست فرمت‌های مجاز پیوست برای پایگاه‌های موجود
    _fmts = get_setting("allowed_formats", "") or ""
    if "csv" not in [e.strip() for e in _fmts.lower().split(",") if e.strip()]:
        set_setting("allowed_formats", (_fmts.rstrip(",") + ",csv").strip(","))
    # عنوان ردیف‌های «ارزیابی امنیتی وب» از فیلد «آدرس» گرفته می‌شود (به‌جای کارفرما)؛
    # پس از تغییر کلید فیلد، عنوان فعالیت‌های موجود این حوزه هم یک‌بار همگام‌سازی می‌شود.
    try:
        _dw = db.execute("SELECT id FROM domains WHERE name='ارزیابی امنیتی وب'"
                         " LIMIT 1").fetchone()
        if _dw:
            _c1 = db.execute("UPDATE form_fields SET field_key=NULL WHERE domain_id=?"
                             " AND label='کارفرما' AND field_key='title'",
                             (_dw["id"],)).rowcount
            _c2 = db.execute("UPDATE form_fields SET field_key='title' WHERE domain_id=?"
                             " AND label='آدرس' AND (field_key IS NULL OR field_key='')",
                             (_dw["id"],)).rowcount
            if _c1 or _c2:
                for _a in db.execute("SELECT id FROM activities WHERE domain_id=?",
                                     (_dw["id"],)).fetchall():
                    _sync_meta(_a["id"])
    except sqlite3.OperationalError:
        pass
    # هم‌ترازی افزودنی حوزه‌ها با نسخهٔ جدید: فیلدهای تعریف‌شدهٔ جدید (مثل «شناسه فرآیند»)
    # به حوزه‌های موجود اضافه می‌شوند — هیچ فیلدی حذف یا تغییرنام نمی‌یابد تا داده‌ها سالم بمانند.
    try:
        for _name, _defs in DOMAIN_FIELDS.items():
            _dom = db.execute("SELECT id FROM domains WHERE name=?", (_name,)).fetchone()
            if not _dom:
                continue
            _have = {r["label"] for r in db.execute(
                "SELECT label FROM form_fields WHERE domain_id=?", (_dom["id"],))}
            _mx = db.execute("SELECT COALESCE(MAX(sort_order),0) m FROM form_fields "
                             "WHERE domain_id=?", (_dom["id"],)).fetchone()["m"]
            for (_label, _ftype, _fkey, _opts, _req, _sec) in _defs:
                if _label in _have:
                    continue
                _mx += 1
                db.execute("INSERT INTO form_fields(domain_id,label,field_key,field_type,"
                           "section,options,required,sort_order) VALUES(?,?,?,?,?,?,?,?)",
                           (_dom["id"], _label, _fkey, _ftype, _sec,
                            json.dumps(_opts, ensure_ascii=False), _req, _mx))
    except sqlite3.OperationalError:
        pass
    db.commit()


# ------------------------------------------------------------------ احراز هویت
def current_user():
    uid = session.get("uid")
    if not uid:
        return None
    return get_db().execute("SELECT * FROM users WHERE id=?", (uid,)).fetchone()


def login_required(fn):
    @wraps(fn)
    def wrapper(*a, **kw):
        user = current_user()
        if not user or not user["is_active"]:
            return redirect(url_for("login", next=request.path))
        g.user = user
        return fn(*a, **kw)
    return wrapper


def admin_required(fn):
    @wraps(fn)
    def wrapper(*a, **kw):
        user = current_user()
        if not user or not user["is_active"]:
            return redirect(url_for("login", next=request.path))
        if user["role"] != "admin":
            abort(403)
        g.user = user
        return fn(*a, **kw)
    return wrapper


PERM_COLS = ("can_add", "can_edit", "can_delete", "can_import")
PERM_NAMES = {"can_add": "ثبت فعالیت", "can_edit": "ویرایش فعالیت",
              "can_delete": "حذف فعالیت", "can_import": "ورود از Excel"}


def perm_required(col):
    """دسترسی بر اساس مجوز دانه‌ای کاربر؛ مدیر همیشه مجاز است."""
    def deco(fn):
        @wraps(fn)
        def wrapper(*a, **kw):
            user = current_user()
            if not user or not user["is_active"]:
                return redirect(url_for("login", next=request.path))
            if user["role"] != "admin" and not user[col]:
                abort(403)
            g.user = user
            return fn(*a, **kw)
        return wrapper
    return deco


_login_attempts = {}   # ip+username → [تعداد ناموفق، زمان اولین تلاش]
_LOGIN_MAX_FAILS, _LOGIN_LOCK_SEC = 5, 300


@app.route("/login", methods=["GET", "POST"])
def login():
    if current_user():
        return redirect(url_for("dashboard"))
    if request.method == "POST":
        import time as _t
        u = request.form.get("username", "").strip()
        p = request.form.get("password", "")
        key = f"{request.remote_addr}|{u.lower()}"
        fails, first = _login_attempts.get(key, [0, 0.0])
        row = get_db().execute("SELECT * FROM users WHERE username=? AND is_active=1",
                               (u,)).fetchone()
        if row and check_password_hash(row["password_hash"], p):
            _login_attempts.pop(key, None)
            session.clear()
            session.permanent = True
            session["uid"] = row["id"]
            return redirect(request.args.get("next") or url_for("dashboard"))
        if fails >= _LOGIN_MAX_FAILS and _t.time() - first < _LOGIN_LOCK_SEC:
            flash("به دلیل تلاش‌های متوالی ناموفق، ورود برای چند دقیقه قفل شده است. "
                  "کمی بعد دوباره تلاش کنید.", "error")
            return render_template("login.html", sys_name=get_setting("system_name")), 429
        _login_attempts[key] = ([fails + 1, first or _t.time()] if _t.time() - first < _LOGIN_LOCK_SEC
                                else [1, _t.time()])
        flash("نام کاربری یا رمز عبور اشتباه است.", "error")
    return render_template("login.html", sys_name=get_setting("system_name"))


@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))


# ------------------------------------------------------------------- ابزارها
def get_domain_or_404(domain_id):
    row = get_db().execute("SELECT * FROM domains WHERE id=?", (domain_id,)).fetchone()
    if not row:
        abort(404)
    return row


def domain_icon(name):
    """شناسه آیکن SVG حوزه (symbol id در اسپرایت)"""
    return DOMAIN_ICONS.get(name, "d-box")


def grouped_fields(fields):
    """گروه‌بندی فیلدها بر اساس بخش: [('', [..]), ('مشخصات ...', [..]), ...]"""
    groups, order = {}, []
    for f in fields:
        sec = f["section"] or ""
        if sec not in groups:
            groups[sec] = []
            order.append(sec)
        groups[sec].append(f)
    return [(s, groups[s]) for s in order]


def field_options(field):
    try:
        return json.loads(field["options"] or "[]")
    except ValueError:
        return []


def _field_json(f):
    """سریالایز فیلد فرم برای رندر سمت کلاینت با Vue 3 (افزودنی)."""
    return {"id": f["id"], "label": f["label"], "type": f["field_type"],
            "options": field_options(f), "required": bool(f["required"]),
            "section": f["section"] or "", "key": f["field_key"] or ""}


def _fields_json(fields):
    return [_field_json(f) for f in fields]


def get_fields(domain_id, only_active=True):
    q = "SELECT * FROM form_fields WHERE domain_id=?"
    if only_active:
        q += " AND is_active=1"
    return get_db().execute(q + " ORDER BY sort_order, id", (domain_id,)).fetchall()


def parse_jalali_triplet(src, prefix, required, label):
    """خواندن سه‌تکه تاریخ شمسی از فرم -> (iso میلادی | None, خطا | None)"""
    y = jalali.to_ascii_digits(src.get(prefix + "_y", "")).strip()
    m = src.get(prefix + "_m", "").strip()
    d = src.get(prefix + "_d", "").strip()
    if not y and not m and not d:
        if required:
            return None, f"فیلد «{label}» الزامی است."
        return None, None
    if not (y and m and d):
        return None, f"تاریخ «{label}» کامل نیست."
    parsed = jalali.j_str_to_g(f"{y}/{m}/{d}")
    if not parsed:
        return None, f"تاریخ «{label}» نامعتبر است."
    return jalali.j_parts_to_g_iso(*parsed), None


def collect_values(fields, form, user, files=None):
    """قرائت مقادیر فیلدهای پویا از فرم -> (dict[field_id]=value, [errors])"""
    values, errors = {}, []
    for f in fields:
        fid, ftype = f["id"], f["field_type"]
        if ftype == "file":
            # فیلدهای فایل: مقدار متنی در فرم نیست؛ فایل جدید یا مقدار قبلی کافی است
            up = (files or {}).get(f"ff{fid}")
            has_old = bool(form.get(f"curf{fid}", "").strip())
            if f["required"] and not (up and up.filename) and not has_old:
                errors.append(f"فیلد «{f['label']}» (فایل) الزامی است.")
            continue
        if ftype == "date":
            iso, err = parse_jalali_triplet(form, f"f{fid}_", f["required"], f["label"])
            if err:
                errors.append(err)
            elif iso:
                values[fid] = iso
        else:
            v = form.get(f"f{fid}", "").strip()
            if ftype == "number" and v:
                v_raw = jalali.to_ascii_digits(v)
                if not re.fullmatch(r"-?\d+(\.\d+)?", v_raw):
                    errors.append(f"فیلد «{f['label']}» باید عددی باشد.")
                    continue
                v = v_raw
            if f["required"] and not v:
                errors.append(f"فیلد «{f['label']}» الزامی است.")
                continue
            values[fid] = v
    return values, errors


def save_values(activity_id, values):
    db = get_db()
    for fid, val in values.items():
        db.execute("INSERT INTO activity_values(activity_id,field_id,value) VALUES(?,?,?)"
                   " ON CONFLICT(activity_id,field_id) DO UPDATE SET value=excluded.value",
                   (activity_id, fid, val))
    _sync_meta(activity_id)


def _sync_meta(activity_id):
    """استخراج عنوان/تاریخ/تیکت از مقادیر پویا برای جستجو و مرتب‌سازی سریع."""
    db = get_db()
    rows = db.execute("""
        SELECT v.value, f.field_key, f.field_type, f.label, f.sort_order
        FROM activity_values v JOIN form_fields f ON f.id=v.field_id
        WHERE v.activity_id=? ORDER BY f.sort_order, f.id""", (activity_id,)).fetchall()
    title, ticket, date_iso, date_fallback = "", "", None, None
    for r in rows:
        val = (r["value"] or "").strip()
        key = r["field_key"] or ""
        if key == "date" and val:
            date_iso = val
        elif r["field_type"] == "date" and val and not date_fallback:
            date_fallback = val
        if not ticket and (key == "ticket" or "تیکت" in r["label"]) and val:
            ticket = jalali.to_ascii_digits(val)
        if not title and (key == "title") and val:
            title = val
    if not date_iso:
        date_iso = date_fallback
    if not title:
        for r in rows:
            if (r["value"] or "").strip() and r["field_type"] in ("text", "textarea") \
                    and (r["field_key"] or "") not in ("expert", "ticket"):
                title = (r["value"] or "").strip().split("\n")[0][:120]
                break
    db.execute("UPDATE activities SET title=?, ticket=?, date=? WHERE id=?",
               (title[:200], ticket[:100], date_iso, activity_id))
    db.commit()


def own_or_admin(activity):
    u = g.user
    return u["role"] == "admin" or activity["user_id"] == u["id"]


def build_filters():
    """فیلترهای مشترک لیست فعالیت‌ها و گزارش‌ها -> (where_sql, params)"""
    u, args = g.user, request.args
    where, params = [], []
    if u["role"] != "admin":
        where.append("a.user_id=?")
        params.append(u["id"])
    dom = args.get("domain", type=int)
    if dom:
        where.append("a.domain_id=?")
        params.append(dom)
    org = args.get("org", type=int)
    if org:
        where.append("a.domain_id IN (SELECT id FROM domains WHERE org_id=?)")
        params.append(org)
    status = args.get("status", "")
    if status in STATUSES:
        where.append("a.status=?")
        params.append(status)
    expert = args.get("expert", type=int)
    if expert and u["role"] == "admin":
        where.append("a.user_id=?")
        params.append(expert)
    ticket = jalali.to_ascii_digits(args.get("ticket", "").strip())
    if ticket:
        ta = jalali.to_ascii_digits(ticket)
        tf = jalali.fa(ta)
        where.append("(a.ticket LIKE ? OR a.ticket LIKE ?)")
        params += [f"%{ta}%", f"%{tf}%"]
    q = args.get("q", "").strip()
    if q:
        where.append("(a.title LIKE ? OR a.ticket LIKE ? OR u.full_name LIKE ? OR "
                     "EXISTS(SELECT 1 FROM activity_values v WHERE v.activity_id=a.id "
                     "AND v.value LIKE ?))")
        params += [f"%{q}%"] * 4
    g_from, err1 = parse_jalali_triplet(args, "from_", False, "از تاریخ")
    g_to, err2 = parse_jalali_triplet(args, "to_", False, "تا تاریخ")
    if g_from:
        where.append("a.date>=?")
        params.append(g_from)
    if g_to:
        where.append("a.date<=?")
        params.append(g_to)
    if args.get("attach") == "1":
        where.append("EXISTS(SELECT 1 FROM attachments t WHERE t.activity_id=a.id)")
    return (" AND ".join(where) or "1=1"), params


def query_activities(where, params, limit=None, offset=0):
    sql = f"""SELECT a.*, d.name AS domain_name, u.full_name AS expert_name,
             (SELECT v.value FROM activity_values v JOIN form_fields ff
               ON ff.id=v.field_id AND ff.field_key='expert'
              WHERE v.activity_id=a.id LIMIT 1) AS expert_txt,
             (SELECT COUNT(*) FROM attachments t WHERE t.activity_id=a.id) AS att_count
             FROM activities a
             JOIN domains d ON d.id=a.domain_id
             JOIN users u ON u.id=a.user_id
             WHERE {where} ORDER BY a.date IS NULL, a.date DESC, a.id DESC"""
    if limit:
        sql += " LIMIT ? OFFSET ?"
        params = params + [limit, offset]
    return get_db().execute(sql, params).fetchall()


def activity_values_map(activity_id):
    rows = get_db().execute("SELECT field_id, value FROM activity_values "
                            "WHERE activity_id=?", (activity_id,)).fetchall()
    return {r["field_id"]: r["value"] for r in rows}


def build_report_charts(acts):
    """نمودارهای خروجی PDF بر اساس همان مجموعه فیلترشده."""
    dom, st, mon = {}, {}, {}
    for a in acts:
        dom[a["domain_name"]] = dom.get(a["domain_name"], 0) + 1
        st[a["status"]] = st.get(a["status"], 0) + 1
        if a["date"]:
            ym = a["date"][:7]
            mon[ym] = mon.get(ym, 0) + 1
    months = []
    for ym in sorted(mon)[-60:]:
        gy, gm = int(ym[:4]), int(ym[5:7])
        jy, jm, _ = jalali.g2j(gy, gm, 15)
        months.append({"label": f"{jalali.MONTH_NAMES[jm]} {jalali.fa(jy)}",
                       "value": mon[ym]})
    return {
        "domains": [{"label": k, "value": v} for k, v in
                    sorted(dom.items(), key=lambda x: -x[1])],
        "status": [{"label": s, "value": st.get(s, 0)} for s in STATUSES],
        "monthly": months,
    }


def filters_summary(module="reports"):
    db, args = get_db(), request.args
    parts = []
    dom = args.get("domain", type=int)
    if dom:
        d = db.execute("SELECT name FROM domains WHERE id=?", (dom,)).fetchone()
        if d:
            parts.append(f"حوزه: {d['name']}")
    org = args.get("org", type=int)
    if org:
        o = db.execute("SELECT name FROM orgs WHERE id=?", (org,)).fetchone()
        if o:
            parts.append(f"مرکز: {o['name']}")
    if args.get("expert") and g.user["role"] == "admin":
        u = db.execute("SELECT full_name FROM users WHERE id=?",
                       (args.get("expert", type=int),)).fetchone()
        if u:
            parts.append(f"کارشناس: {u['full_name']}")
    if g.user["role"] != "admin":
        parts.append(f"کارشناس: {g.user['full_name']} (فقط فعالیت‌های خود)")
    g_from, _ = parse_jalali_triplet(args, "from_", False, "از تاریخ")
    g_to, _ = parse_jalali_triplet(args, "to_", False, "تا تاریخ")
    rng = []
    if g_from:
        rng.append("از " + jalali.fa(jalali.g_str_to_j(g_from)))
    if g_to:
        rng.append("تا " + jalali.fa(jalali.g_str_to_j(g_to)))
    if rng:
        parts.append("بازه زمانی: " + " ".join(rng))
    if args.get("status") in STATUSES:
        parts.append("وضعیت: " + args["status"])
    if args.get("ticket"):
        parts.append("تیکت: " + jalali.fa(args["ticket"]))
    if args.get("attach") == "1":
        parts.append("فقط دارای ضمیمه")
    return " • ".join(parts) if parts else "بدون فیلتر (همه موارد)"


BASE_COLS = ["حوزه", "عنوان", "کارشناس", "وضعیت", "تاریخ (شمسی)", "شماره تیکت"]
SKIP_FIELD_LABELS = {"کارشناس", "تاریخ", "وضعیت", "شماره تیکت"}


def export_rows(acts):
    """خروجی = همه فیلدهای داخل فعالیت‌ها (بدون شناسه سیستمی، بدون ستون تکراری)."""
    db = get_db()
    ids = [a["id"] for a in acts]
    seen_labels, field_types, fields_map = [], {}, {}
    if ids:
        # پیش‌بذر ستون‌ها از تعریف فیلدهای فعالِ حوزه‌های موجود در نتیجه — تا فیلدی
        # که تازه به حوزه اضافه شده و هنوز روی هیچ فعالیتی مقدار ندارد هم ستون بگیرد.
        dom_ids = sorted({a["domain_id"] for a in acts if "domain_id" in a.keys()})
        if dom_ids:
            phd = ",".join("?" * len(dom_ids))
            for df in db.execute(f"""SELECT label FROM form_fields
                                     WHERE domain_id IN ({phd}) AND is_active=1
                                     ORDER BY domain_id, sort_order""", dom_ids):
                if (df["label"] not in seen_labels
                        and df["label"] not in SKIP_FIELD_LABELS
                        and df["label"] not in BASE_COLS):
                    seen_labels.append(df["label"])
        ph = ",".join("?" * len(ids))
        rows = db.execute(f"""SELECT v.activity_id, v.field_id, v.value, f.label, f.field_type
                              FROM activity_values v JOIN form_fields f ON f.id=v.field_id
                              WHERE v.activity_id IN ({ph}) ORDER BY f.sort_order""",
                          ids).fetchall()
        for r in rows:
            if r["label"] not in seen_labels and r["label"] not in SKIP_FIELD_LABELS                     and r["label"] not in BASE_COLS:
                seen_labels.append(r["label"])
            field_types.setdefault(r["label"], r["field_type"])
            fields_map.setdefault(r["activity_id"], {})[r["label"]] = (
                r["value"], r["field_type"])
    header = list(BASE_COLS) + seen_labels
    data = []
    for a in acts:
        vals = fields_map.get(a["id"], {})
        row = [a["domain_name"], a["title"] or "",
               _canon_expert((a["expert_txt"] or a["expert_name"]) if "expert_txt" in a.keys() else a["expert_name"]),
               a["status"],
               jalali.fa(jalali.g_str_to_j(a["date"])) if a["date"] else "",
               a["ticket"] or ""]
        for c in seen_labels:
            val_typ = vals.get(c)
            if not val_typ:
                row.append("")
                continue
            v, t = val_typ
            if t == "date" and v:
                v = jalali.fa(jalali.g_str_to_j(v))
            row.append("" if v is None else v)
        data.append(row)
    # ستون «شماره تیکت» فقط وقتی می‌آید که حداقل یک ردیف تیکت داشته باشد
    # (برخی حوزه‌ها مثل ارزیابی امنیتی وب اصلاً تیکت ندارند)
    if data and "شماره تیکت" in header and all(not r[header.index("شماره تیکت")] for r in data):
        ti = header.index("شماره تیکت")
        header.pop(ti)
        for r in data:
            r.pop(ti)
    return header, data


def select_cols(header, rows, cols):
    """فیلتر کردن ستون‌های خروجی بر اساس انتخاب کاربر."""
    idx = [i for i, h in enumerate(header) if h in cols]
    if not idx:
        return header, rows
    return [header[i] for i in idx], [[r[i] for i in idx] for r in rows]


def export_cols_param():
    if g.user["role"] == "admin":
        return request.args.getlist("col")
    return []


# ------------------------------------------------------------------- داشبورد
def dashboard_payload():
    """محاسبه داده‌های داشبورد؛ خروجی: (متغیرهای قالب، جیسون API/Vue)"""
    db, u = get_db(), g.user
    mine = "" if u["role"] == "admin" else "AND a.user_id=?"
    mp = [] if u["role"] == "admin" else [u["id"]]

    per_domain = db.execute(f"""
        SELECT d.name, COUNT(a.id) c FROM domains d
        LEFT JOIN activities a ON a.domain_id=d.id {mine}
        WHERE d.is_active=1 GROUP BY d.id ORDER BY c DESC""", mp).fetchall()
    per_status = db.execute(f"""
        SELECT a.status, COUNT(*) c FROM activities a WHERE 1=1 {mine}
        GROUP BY a.status""", mp).fetchall()
    daily = db.execute(f"""
        SELECT substr(a.date,1,10) d, COUNT(*) c FROM activities a
        WHERE a.date IS NOT NULL {mine} GROUP BY d""",
                         mp).fetchall()
    total = db.execute(f"SELECT COUNT(*) c FROM activities a WHERE 1=1 {mine}",
                       mp).fetchone()["c"]
    last_dt = db.execute(f"""SELECT MAX(a.date) m FROM activities a
                             WHERE a.date IS NOT NULL {mine}""",
                         mp).fetchone()["m"]
    # سبد ماه شمسی: تجمیع روزها بر مرز ماه شمسی (نه میلادی) تا ماه‌ها دقیق شوند
    m_counts = {}
    for row in daily:
        d = row["d"] or ""
        if len(d) < 10:
            continue
        try:
            jy, jm, _ = jalali.g2j(int(d[:4]), int(d[5:7]), int(d[8:10]))
        except ValueError:
            continue
        k = f"{jy:04d}-{jm:02d}"
        m_counts[k] = m_counts.get(k, 0) + row["c"]
    tj = jalali.today_jalali()
    this_month = m_counts.get(f"{tj[0]:04d}-{tj[1]:02d}", 0)
    # دلتای رشد نسبت به ماه شمسی قبل
    _py, _pm = (tj[0], tj[1] - 1) if tj[1] > 1 else (tj[0] - 1, 12)
    prev_month = m_counts.get(f"{_py:04d}-{_pm:02d}", 0)
    month_delta = this_month - prev_month
    users_c = db.execute("SELECT COUNT(*) c FROM users WHERE is_active=1").fetchone()["c"]
    last_uploads = db.execute("""
        SELECT e.*, d.name domain_name, u.full_name FROM excel_imports e
        LEFT JOIN domains d ON d.id=e.domain_id LEFT JOIN users u ON u.id=e.user_id
        ORDER BY e.id DESC LIMIT 5""").fetchall()
    # تفکیک کارشناس بر اساس «نام نمایشی»: همان نامی که در فیلد کارشناسِ فعالیت ثبت
    # شده (حتی اگر در فهرست کاربران سامانه نباشد)؛ نام مدیران سامانه حذف می‌شود.
    per_expert = db.execute(f"""
        SELECT n, COUNT(*) c FROM (
            SELECT COALESCE(NULLIF(TRIM((SELECT v.value FROM activity_values v
                     JOIN form_fields ff ON ff.id=v.field_id AND ff.field_key='expert'
                     WHERE v.activity_id=a.id LIMIT 1)), ''), u.full_name) AS n
            FROM activities a JOIN users u ON u.id=a.user_id WHERE 1=1 {mine}
        ) GROUP BY n""", mp).fetchall()
    # ادغام نام‌های ناهماهنگ («رضایی» با «علی رضایی») پیش از ترسیم نمودار
    _admin_names = {r["full_name"] for r in
                    db.execute("SELECT full_name FROM users WHERE role='admin'")}
    _mrg = {}
    for _r in per_expert:
        _cn = _canon_expert(_r["n"])
        if _cn in _admin_names:
            continue
        _mrg[_cn] = _mrg.get(_cn, 0) + _r["c"]
    per_expert = sorted(_mrg.items(), key=lambda kv: (-kv[1], kv[0]))[:15]
    last_acts = db.execute(f"""
        SELECT a.*, d.name domain_name, u.full_name expert_name FROM activities a
        JOIN domains d ON d.id=a.domain_id JOIN users u ON u.id=a.user_id
        WHERE 1=1 {mine} ORDER BY a.id DESC LIMIT 5""", mp).fetchall()

    # بازه پیوسته ماه‌های شمسی: از اولین ماه دارای داده تا ماه جاری (خالی‌ها با صفر)
    if m_counts:
        _first = min(m_counts)
        _y, _m = int(_first[:4]), int(_first[5:7])
    else:
        _y, _m = tj[0], tj[1]
    months = []
    while (_y, _m) <= (tj[0], tj[1]):
        months.append({"label": f"{jalali.MONTH_NAMES[_m]} {jalali.fa(_y)}",
                       "value": m_counts.get(f"{_y:04d}-{_m:02d}", 0)})
        _m += 1
        if _m > 12:
            _y, _m = _y + 1, 1
    months = months[-60:]  # تا ۵ سال تاریخچه — بازه با فیلتر سمت کلاینت محدود می‌شود
    charts = {
        "domains": [{"label": r["name"], "value": r["c"]} for r in per_domain],
        "status": [{"label": s, "value": next((r["c"] for r in per_status
                                               if r["status"] == s), 0)}
                   for s in STATUSES],
        "monthly": months,
        "experts": [{"label": n, "value": c} for n, c in per_expert],
    }
    today = jalali.today_jalali()
    tv = {"charts": charts, "total": total, "this_month": this_month,
          "users_c": users_c, "month_delta": month_delta, "prev_month": prev_month,
          "last_uploads": last_uploads, "last_acts": last_acts,
          "last_date_fa": _jdate(last_dt) if last_dt else "",
          "today_fa": jalali.fa(f"{today[0]:04d}/{today[1]:02d}/{today[2]:02d}")}
    payload = {
        "total": total, "this_month": this_month, "month_delta": month_delta,
        "last_date": tv["last_date_fa"],
        "domains_n": len(charts["domains"]), "users_c": users_c, "charts": charts,
        "last_acts": [{"id": a["id"], "title": a["title"] or "بدون عنوان",
                       "domain": a["domain_name"], "icon": domain_icon(a["domain_name"]),
                       "status": a["status"], "ticket": a["ticket"] or "",
                       "date": _jdate(a["date"]) if a["date"] else "—",
                       "view": url_for("activity_view", activity_id=a["id"])}
                      for a in last_acts],
        "last_uploads": [{"file": e["filename"], "domain": e["domain_name"] or "—",
                          "user": e["full_name"] or "—", "ok": e["success_rows"],
                          "err": e["error_rows"], "when": _jdatetime(e["imported_at"])}
                         for e in last_uploads],
        "today_fa": tv["today_fa"],
    }
    return tv, payload


@app.route("/")
@login_required
def dashboard():
    tv, payload = dashboard_payload()
    return render_template("dashboard.html", dash_payload=payload, **tv)


@app.route("/api/dashboard")
@login_required
def api_dashboard():
    """داده به‌روز داشبورد برای رفرش زنده Vue — فقط افزودنی، خواندنی."""
    _, payload = dashboard_payload()
    return jsonify(payload)


# ----------------------------------------------------------------- فعالیت‌ها
@app.route("/activities")
@login_required
def activities():
    where, params = build_filters()
    total = get_db().execute(f"""SELECT COUNT(*) c FROM activities a
             JOIN users u ON u.id=a.user_id WHERE {where}""", params).fetchone()["c"]

    if request.args.get("export") == "excel":
        return export_excel(query_activities(where, params))
    if request.args.get("export") == "pdf":
        all_acts = query_activities(where, params)
        header, rows = export_rows(all_acts)
        return render_template("report_print.html", header=header, rows=rows,
                               title="گزارش فعالیت‌ها", charts=build_report_charts(all_acts),
                               filter_text=filters_summary(),
                               sys_name=get_setting("system_name"), autoprint=True)

    # همه ردیف‌های مطابق فیلتر برای جدول تعاملی Vue (جستجو/مرتب‌سازی/صفحه‌بندی سمت کلاینت)
    acts = query_activities(where, params)
    domains = get_db().execute("SELECT * FROM domains WHERE is_active=1 "
                               "ORDER BY sort_order, id").fetchall()
    users = get_db().execute("SELECT id,full_name FROM users WHERE is_active=1 "
                             "ORDER BY full_name").fetchall()
    # داده‌های JSON برای رندر سمت کلاینت با Vue 3 (افزودنی)
    acts_json = [{
        "id": a["id"],
        "title": a["title"] or "بدون عنوان",
        "domain": a["domain_name"],
        "icon": domain_icon(a["domain_name"]),
        "expert": _canon_expert(a["expert_txt"] or a["expert_name"]),
        "flagged": a["flagged"] if "flagged" in a.keys() else 0,
        "status": a["status"],
        "date": _jdate(a["date"]) if a["date"] else "—",
        "date_key": a["date"] or "",
        "ticket": a["ticket"] or "",
        "atts": a["att_count"] or 0,
        "task": bool(a["created_by"] and a["created_by"] != a["user_id"]),
        "view": url_for("activity_view", activity_id=a["id"]),
        "edit": url_for("activity_edit", activity_id=a["id"]),
        "delete": url_for("activity_delete", activity_id=a["id"]),
    } for a in acts]
    flt_domains = [{"id": d["id"], "name": d["name"], "org_id": d["org_id"] or 0}
                   for d in domains]
    flt_users = [{"id": u["id"], "full_name": u["full_name"]} for u in users]
    orgs = get_db().execute("SELECT * FROM orgs ORDER BY sort_order, id").fetchall()
    return render_template("activities.html", acts=acts, acts_json=acts_json,
                           total=total, domains=domains, flt_domains=flt_domains,
                           flt_users=flt_users, orgs=orgs,
                           orgs_json=[{"id": o["id"], "name": o["name"]} for o in orgs],
                           bulk_statuses=(STATUSES if g.user["role"] == "admin"
                                          else EXPERT_STATUSES),
                           users=users, statuses=STATUSES)


@app.route("/activities/new", methods=["GET", "POST"])
@perm_required("can_add")
def activity_new():
    db = get_db()
    domain_id = request.values.get("domain_id", type=int)
    domains = db.execute("""SELECT d.*, o.name org_name FROM domains d
                            LEFT JOIN orgs o ON o.id=d.org_id
                            WHERE d.is_active=1
                            ORDER BY COALESCE(o.sort_order,99), d.sort_order, d.id""").fetchall()
    if not domain_id:
        return render_template(
            "activity_new.html", domains=domains, domain=None, icons=DOMAIN_ICONS,
            domains_json=[{"id": d["id"], "name": d["name"], "org": d["org_name"] or "سایر",
                           "icon": domain_icon(d["name"])} for d in domains])
    domain = get_domain_or_404(domain_id)
    if not domain["is_active"]:
        abort(404)
    fields = get_fields(domain_id)
    if request.method == "POST":
        values, errors = collect_values(fields, request.form, g.user, request.files)
        status = request.form.get("status", STATUSES[0])
        if status not in STATUSES:
            status = STATUSES[0]
        if g.user["role"] != "admin" and status == "بررسی شده":
            status = STATUSES[1]
        if errors:
            for e in errors:
                flash(e, "error")
        else:
            owner_id = g.user["id"]
            if g.user["role"] == "admin":
                oid = request.form.get("owner_id", type=int)
                if oid and db.execute("SELECT 1 FROM users WHERE id=? AND is_active=1",
                                      (oid,)).fetchone():
                    owner_id = oid
            cur = db.execute("INSERT INTO activities(domain_id,user_id,status,"
                             "created_at,updated_at,created_by) VALUES(?,?,?,?,?,?)",
                             (domain_id, owner_id, status, now_iso(), now_iso(),
                              g.user["id"]))
            save_values(cur.lastrowid, values)
            _save_field_files(db, cur.lastrowid, fields, request.files)
            db.commit()
            owner = db.execute("SELECT full_name FROM users WHERE id=?",
                               (owner_id,)).fetchone()
            flash("فعالیت با موفقیت ثبت شد" +
                  (f" و به «{owner['full_name']}» تخصیص یافت."
                   if owner_id != g.user["id"] else "."), "success")
            return redirect(url_for("activity_view", activity_id=cur.lastrowid))
    defaults, dvals = {}, {}
    for f in fields:
        if f["field_key"] == "expert" and not request.form.get(f"f{f['id']}"):
            defaults[f["id"]] = g.user["full_name"]
        if f["field_type"] == "date":
            y, m, d = jalali.today_jalali()
            posted = {k: request.form.get(f"f{f['id']}__{k}", "")
                      for k in ("y", "m", "d")}
            dvals[f["id"]] = {k: posted[k] or v for k, v in
                              (("y", str(y)), ("m", f"{m:02d}"), ("d", f"{d:02d}"))}
    users = []
    if g.user["role"] == "admin":
        users = db.execute("SELECT id, full_name FROM users WHERE is_active=1 "
                           "ORDER BY full_name").fetchall()
    return render_template("activity_new.html", domains=domains, domain=domain,
                           fields=fields, groups=grouped_fields(fields),
                           defaults=defaults, dvals=dvals, users=users,
                           fields_json=_fields_json(fields),
                           posted=dict(request.form),
                           users_json=[{"id": x["id"], "full_name": x["full_name"]}
                                       for x in users],
                           domains_json=[{"id": d["id"], "name": d["name"],
                                          "icon": domain_icon(d["name"])}
                                         for d in domains],
                           statuses=(STATUSES if g.user["role"] == "admin"
                                     else EXPERT_STATUSES), icons=DOMAIN_ICONS)


def get_activity_or_404(activity_id):
    a = get_db().execute("""SELECT a.*, d.name domain_name, u.full_name expert_name
                            FROM activities a JOIN domains d ON d.id=a.domain_id
                            JOIN users u ON u.id=a.user_id WHERE a.id=?""",
                         (activity_id,)).fetchone()
    if not a:
        abort(404)
    return a


@app.route("/activities/<int:activity_id>")
@login_required
def activity_view(activity_id):
    a = get_activity_or_404(activity_id)
    if not own_or_admin(a):
        abort(403)
    fields = get_fields(a["domain_id"])
    vals = activity_values_map(a["id"])
    db = get_db()
    atts = db.execute("SELECT * FROM attachments WHERE activity_id=? "
                      "ORDER BY id", (activity_id,)).fetchall()
    responses = db.execute("""SELECT r.*, u.full_name, u.role FROM responses r
                              JOIN users u ON u.id=r.user_id WHERE r.activity_id=?
                              ORDER BY r.id""", (activity_id,)).fetchall()
    creator = None
    if a["created_by"]:
        creator = db.execute("SELECT full_name FROM users WHERE id=?",
                             (a["created_by"],)).fetchone()
    # داده JSON برای رندر تعاملی Vue 3 (افزودنی)
    att_by_name = {t["original_name"]: url_for("attachment_download", att_id=t["id"])
                   for t in atts}
    vals_sections = []
    for sec, fs in grouped_fields(fields):
        vals_sections.append({
            "name": sec or "",
            "icon": "i-user" if "درخواست‌دهنده" in sec else "i-send",
            "rows": [{"label": f["label"],
                      "ftype": f["field_type"],
                      "val": (_jdate(vals[f["id"]]) if f["field_type"] == "date"
                              and vals.get(f["id"]) else (vals.get(f["id"]) or "")),
                      "dl": (att_by_name.get(vals.get(f["id"]) or "", "")
                             if f["field_type"] == "file" else "")}
                     for f in fs],
        })
    view_json = {
        "title": a["title"] or "بدون عنوان",
        "domain": a["domain_name"], "icon": domain_icon(a["domain_name"]),
        "status": a["status"], "expert": _canon_expert((a["expert_txt"] or a["expert_name"]) if "expert_txt" in a.keys() else a["expert_name"]),
        "creator": creator["full_name"] if creator else "",
        "flagged": bool(a["flagged"]) if "flagged" in a.keys() else False,
        "date_fa": _jdate(a["date"]) if a["date"] else "",
        "ticket": a["ticket"] or "",
        "back_url": url_for("activities"),
        "vals_sections": vals_sections,
        "responses": [{"user": r["full_name"], "role": r["role"],
                       "when": _jdatetime(r["created_at"]), "body": r["body"]}
                      for r in responses],
        "atts": [{"name": t["original_name"],
                  "kb": int((t["size"] or 0) / 1024),
                  "when": _jdatetime(t["uploaded_at"]),
                  "download": url_for("attachment_download", att_id=t["id"]),
                  "delete": url_for("attachment_delete", att_id=t["id"]),
                  "can_del": g.user["role"] == "admin" or t["uploaded_by"] == g.user["id"]}
                 for t in atts],
        "respond_action": url_for("activity_respond", activity_id=a["id"]),
        "upload_action": url_for("attachment_upload", activity_id=a["id"]),
        "respond_hint": bool(a["created_by"] and a["created_by"] != a["user_id"]
                             and g.user["id"] == a["user_id"]),
    }
    return render_template("activity_view.html", a=a, fields=fields, creator=creator,
                           groups=grouped_fields(fields), vals=vals, atts=atts,
                           responses=responses, view_json=view_json)


@app.route("/activities/<int:activity_id>/respond", methods=["POST"])
@login_required
def activity_respond(activity_id):
    a = get_activity_or_404(activity_id)
    if not own_or_admin(a):
        abort(403)
    body = request.form.get("body", "").strip()
    if not body:
        flash("متن پاسخ خالی است.", "error")
        return redirect(url_for("activity_view", activity_id=activity_id))
    get_db().execute("INSERT INTO responses(activity_id,user_id,body,created_at) "
                     "VALUES(?,?,?,?)", (activity_id, g.user["id"], body, now_iso()))
    get_db().commit()
    flash("پاسخ ثبت شد.", "success")
    return redirect(url_for("activity_view", activity_id=activity_id))


@app.route("/activities/<int:activity_id>/complete", methods=["POST"])
@login_required
def activity_complete(activity_id):
    a = get_activity_or_404(activity_id)
    if not own_or_admin(a):
        abort(403)
    get_db().execute("UPDATE activities SET status=?, updated_at=? WHERE id=?",
                     (STATUSES[1], now_iso(), activity_id))
    get_db().commit()
    flash("فعالیت «انجام شده» علامت‌گذاری شد. ✅", "success")
    return redirect(url_for("activity_view", activity_id=activity_id))


@app.route("/activities/<int:activity_id>/edit", methods=["GET", "POST"])
@perm_required("can_edit")
def activity_edit(activity_id):
    a = get_activity_or_404(activity_id)
    if not own_or_admin(a):
        abort(403)
    fields = get_fields(a["domain_id"])
    vals = activity_values_map(a["id"])
    if request.method == "POST":
        values, errors = collect_values(fields, request.form, g.user, request.files)
        status = request.form.get("status", a["status"])
        if status not in STATUSES:
            status = a["status"]
        if g.user["role"] != "admin" and status == "بررسی شده":
            status = a["status"] if a["status"] == "بررسی شده" else STATUSES[1]
        if errors:
            for e in errors:
                flash(e, "error")
        else:
            db = get_db()
            owner_id = a["user_id"]
            if g.user["role"] == "admin":
                oid = request.form.get("owner_id", type=int)
                if oid and db.execute("SELECT 1 FROM users WHERE id=? AND is_active=1",
                                      (oid,)).fetchone():
                    owner_id = oid
            db.execute("UPDATE activities SET status=?, user_id=?, updated_at=?, flagged=0 WHERE id=?",
                       (status, owner_id, now_iso(), activity_id))
            save_values(activity_id, values)
            _save_field_files(db, activity_id, fields, request.files)
            db.commit()
            flash("فعالیت به‌روزرسانی شد.", "success")
            return redirect(url_for("activity_view", activity_id=activity_id))
    dvals = {}
    for f in fields:
        if f["field_type"] == "date":
            v = vals.get(f["id"], "")
            if v:
                jy, jm, jd = jalali.g2j(*[int(x) for x in v[:10].split("-")])
                cur = {"y": str(jy), "m": f"{jm:02d}", "d": f"{jd:02d}"}
            else:
                cur = {"y": "", "m": "", "d": ""}
            dvals[f["id"]] = {k: request.form.get(f"f{f['id']}__{k}", cur[k])
                              for k in ("y", "m", "d")}
    atts = get_db().execute("SELECT * FROM attachments WHERE activity_id=? "
                            "ORDER BY id", (activity_id,)).fetchall()
    users = []
    if g.user["role"] == "admin":
        users = get_db().execute("SELECT id, full_name FROM users WHERE is_active=1 "
                                 "ORDER BY full_name").fetchall()
    return render_template(
        "activity_edit.html", a=a, fields=fields, groups=grouped_fields(fields),
        vals=vals, dvals=dvals, atts=atts, users=users,
        fields_json=_fields_json(fields), posted=dict(request.form),
        users_json=[{"id": x["id"], "full_name": x["full_name"]} for x in users],
        statuses=(STATUSES if g.user["role"] == "admin"
                  else ([s for s in EXPERT_STATUSES] if a["status"] != "بررسی شده"
                        else STATUSES)))


@app.route("/activities/bulk-delete", methods=["POST"])
@perm_required("can_delete")
def activities_bulk_delete():
    """حذف گروهی فعالیت‌ها — بدنه: {"ids": [..]}"""
    ids = request.get_json(silent=True) or {}
    ids = ids.get("ids") or []
    ids = [int(x) for x in ids if str(x).isdigit()][:500]
    db = get_db()
    n = 0
    for aid in ids:
        row = db.execute("SELECT id, user_id FROM activities WHERE id=?",
                         (aid,)).fetchone()
        if not row:
            continue
        if g.user["role"] != "admin" and row["user_id"] != g.user["id"]:
            continue  # کارشناس فقط فعالیت‌های خودش را می‌تواند حذف کند
        for t in db.execute("SELECT stored_name FROM attachments WHERE activity_id=?",
                            (aid,)):
            path = os.path.join(UPLOAD_DIR, t["stored_name"])
            if os.path.exists(path):
                os.remove(path)
        db.execute("DELETE FROM responses WHERE activity_id=?", (aid,))
        db.execute("DELETE FROM activities WHERE id=?", (aid,))
        n += 1
    db.commit()
    return jsonify({"ok": True, "deleted": n})


@app.route("/activities/bulk-status", methods=["POST"])
@perm_required("can_edit")
def activities_bulk_status():
    """تغییر وضعیت گروهی — بدنه: {"ids": [..], "status": "انجام شده"}"""
    body = request.get_json(silent=True) or {}
    ids = [int(x) for x in (body.get("ids") or []) if str(x).isdigit()][:500]
    status = str(body.get("status") or "").strip()
    # کارشناس نمی‌تواند «بررسی شده» بزند (مثل ثبت/ویرایش تکی)
    allowed = STATUSES if g.user["role"] == "admin" else EXPERT_STATUSES
    if status not in allowed:
        return jsonify({"ok": False, "error": "وضعیت نامعتبر است"}), 400
    db = get_db()
    n = 0
    for aid in ids:
        row = db.execute("SELECT id, user_id FROM activities WHERE id=?",
                         (aid,)).fetchone()
        if not row:
            continue
        if g.user["role"] != "admin" and row["user_id"] != g.user["id"]:
            continue  # کارشناس فقط روی فعالیت‌های خودش
        db.execute("UPDATE activities SET status=?, flagged=0, updated_at=? WHERE id=?",
                   (status, now_iso(), aid))
        n += 1
    db.commit()
    return jsonify({"ok": True, "changed": n})


@app.route("/activities/<int:activity_id>/delete", methods=["POST"])
@perm_required("can_delete")
def activity_delete(activity_id):
    a = get_activity_or_404(activity_id)
    db = get_db()
    for t in db.execute("SELECT stored_name FROM attachments WHERE activity_id=?",
                        (activity_id,)):
        path = os.path.join(UPLOAD_DIR, t["stored_name"])
        if os.path.exists(path):
            os.remove(path)
    db.execute("DELETE FROM responses WHERE activity_id=?", (activity_id,))
    db.execute("DELETE FROM activities WHERE id=?", (activity_id,))
    db.commit()
    flash(f"فعالیت «{a['title'] or a['id']}» حذف شد.", "success")
    return redirect(url_for("activities"))


# --------------------------------------------------------------------- تخصیص تسک
def _save_attachments(db, activity_id, files):
    """ذخیره فایل‌های پیوست یک فعالیت (افزودنی — بدون تغییر رفتار قبلی)."""
    saved, errs = 0, []
    os.makedirs(UPLOAD_DIR, exist_ok=True)
    for f in files:
        if not f or not f.filename:
            continue
        ext = f.filename.rsplit(".", 1)[-1].lower() if "." in f.filename else ""
        if ext not in allowed_formats():
            errs.append(f"فرمت «{f.filename}» مجاز نیست.")
            continue
        data = f.read()
        if len(data) > max_upload_mb() * 1024 * 1024:
            errs.append(f"حجم «{f.filename}» بیش از حد مجاز ({max_upload_mb()}MB) است.")
            continue
        stored = f"{secrets.token_hex(10)}.{ext}"
        with open(os.path.join(UPLOAD_DIR, stored), "wb") as out:
            out.write(data)
        db.execute("INSERT INTO attachments(activity_id,stored_name,original_name,size,"
                   "uploaded_by,uploaded_at) VALUES(?,?,?,?,?,?)",
                   (activity_id, stored, f.filename, len(data), g.user["id"], now_iso()))
        saved += 1
    return saved, errs


def _save_field_files(db, activity_id, fields, files):
    """ذخیره فایل فیلدهای «file» فرم پویا: فایل به پیوست‌ها می‌رود و مقدار فیلد = نام فایل."""
    if not files:
        return
    for f in fields:
        if f["field_type"] != "file":
            continue
        up = files.get(f"ff{f['id']}")
        if not up or not up.filename:
            continue
        saved, errs = _save_attachments(db, activity_id, [up])
        for e in errs:
            flash(e, "error")
        if saved:
            db.execute("INSERT INTO activity_values(activity_id,field_id,value) VALUES(?,?,?)"
                       " ON CONFLICT(activity_id,field_id) DO UPDATE SET value=excluded.value",
                       (activity_id, f["id"], up.filename))
    _sync_meta(activity_id)


@app.route("/tasks", methods=["GET", "POST"])
@login_required
def tasks_page():
    db = get_db()
    if request.method == "POST":
        if g.user["role"] != "admin":
            abort(403)
        expert_id = request.form.get("expert_id", type=int)
        domain_id = request.form.get("domain_id", type=int)
        title = (request.form.get("title") or "").strip()
        ticket = (request.form.get("ticket") or "").strip() or None
        note = (request.form.get("note") or "").strip()
        expert = db.execute("SELECT id, full_name FROM users WHERE id=? AND is_active=1",
                            (expert_id,)).fetchone() if expert_id else None
        errors = []
        if not expert:
            errors.append("کارشناس معتبر انتخاب کنید.")
        if not (domain_id and db.execute("SELECT 1 FROM domains WHERE id=? AND is_active=1",
                                         (domain_id,)).fetchone()):
            errors.append("حوزه معتبر انتخاب کنید.")
        if not title:
            errors.append("عنوان تسک را بنویسید.")
        if errors:
            for e in errors:
                flash(e, "error")
        else:
            cur = db.execute(
                "INSERT INTO activities(domain_id,user_id,status,title,ticket,created_at,"
                "updated_at,created_by,task_note) VALUES(?,?,?,?,?,?,?,?,?)",
                (domain_id, expert["id"], STATUSES[0], title, ticket, now_iso(), now_iso(),
                 g.user["id"], note))
            saved, ferrs = _save_attachments(db, cur.lastrowid,
                                             request.files.getlist("files"))
            db.commit()
            msg = f"تسک «{title}» به «{expert['full_name']}» تخصیص یافت."
            if saved:
                msg += f" ({jalali.fa(saved)} فایل پیوست شد)"
            flash(msg, "success")
            for e in ferrs:
                flash(e, "error")
            return redirect(url_for("tasks_page"))
    base_sql = """SELECT a.*, d.name domain_name, u.full_name expert_name,
                     cr.full_name creator_name,
                     (SELECT COUNT(*) FROM attachments t WHERE t.activity_id=a.id) att_c,
                     (SELECT COUNT(*) FROM responses r WHERE r.activity_id=a.id) resp_c
                  FROM activities a JOIN domains d ON d.id=a.domain_id
                  JOIN users u ON u.id=a.user_id
                  LEFT JOIN users cr ON cr.id=a.created_by
                  WHERE a.created_by IS NOT NULL AND a.created_by != a.user_id"""
    users, domains, rows = [], [], []
    if g.user["role"] == "admin":
        rows = db.execute(base_sql + " ORDER BY a.created_at DESC, a.id DESC").fetchall()
        users = db.execute("SELECT id, full_name FROM users WHERE is_active=1 AND id!=? "
                           "ORDER BY full_name", (g.user["id"],)).fetchall()
        domains = db.execute("SELECT id, name FROM domains WHERE is_active=1 "
                             "ORDER BY name").fetchall()
    else:
        rows = db.execute(base_sql + " AND a.user_id=? ORDER BY a.created_at DESC, "
                          "a.id DESC", (g.user["id"],)).fetchall()
    rows_json = [{
        "id": a["id"],
        "who": a["expert_name"] if g.user["role"] == "admin" else (a["creator_name"] or "—"),
        "domain": a["domain_name"],
        "icon": domain_icon(a["domain_name"]),
        "title": a["title"] or "—",
        "ticket": a["ticket"] or "",
        "note": a["task_note"] or "",
        "note_short": ((a["task_note"] or "")[:70]
                       + ("…" if len(a["task_note"] or "") > 70 else "")),
        "status": a["status"],
        "date": _jdate(a["created_at"][:10]),
        "date_key": a["created_at"] or "",
        "atts": a["att_c"] or 0,
        "resps": a["resp_c"] or 0,
        "view": url_for("activity_view", activity_id=a["id"]),
    } for a in rows]
    return render_template("tasks.html", rows=rows, rows_json=rows_json,
                           users=users, domains=domains)


# --------------------------------------------------------------------- پیوست‌ها
@app.route("/activities/<int:activity_id>/attachments", methods=["POST"])
@perm_required("can_edit")
def attachment_upload(activity_id):
    a = get_activity_or_404(activity_id)
    if not own_or_admin(a):
        abort(403)
    files = request.files.getlist("files")
    if not files or all(not f.filename for f in files):
        flash("فایلی انتخاب نشده است.", "error")
        return redirect(request.referrer or url_for("activity_view", activity_id=a["id"]))
    db = get_db()
    saved, errs = _save_attachments(db, activity_id, files)
    db.commit()
    if saved:
        flash(f"{jalali.fa(saved)} فایل پیوست شد.", "success")
    for e in errs:
        flash(e, "error")
    return redirect(request.referrer or url_for("activity_view", activity_id=a["id"]))


@app.route("/attachments/<int:att_id>/download")
@login_required
def attachment_download(att_id):
    db = get_db()
    t = db.execute("SELECT * FROM attachments WHERE id=?", (att_id,)).fetchone()
    if not t:
        abort(404)
    a = get_activity_or_404(t["activity_id"])
    if not own_or_admin(a):
        abort(403)
    path = os.path.join(UPLOAD_DIR, t["stored_name"])
    if not os.path.exists(path):
        abort(404)
    return send_file(path, as_attachment=True, download_name=t["original_name"])


@app.route("/attachments/<int:att_id>/delete", methods=["POST"])
@perm_required("can_edit")
def attachment_delete(att_id):
    db = get_db()
    t = db.execute("SELECT * FROM attachments WHERE id=?", (att_id,)).fetchone()
    if not t:
        abort(404)
    a = get_activity_or_404(t["activity_id"])
    if not (g.user["role"] == "admin" or t["uploaded_by"] == g.user["id"]):
        abort(403)
    path = os.path.join(UPLOAD_DIR, t["stored_name"])
    if os.path.exists(path):
        os.remove(path)
    db.execute("DELETE FROM attachments WHERE id=?", (att_id,))
    db.commit()
    flash("فایل حذف شد.", "success")
    return redirect(request.referrer or url_for("activity_view", activity_id=a["id"]))


# --------------------------------------------------------------- خروجی Excel
def export_excel(acts, filename="activities.xlsx", title="گزارش فعالیت‌ها",
                 override=None):
    if Workbook is None:
        flash("کتابخانه openpyxl نصب نیست.", "error")
        return redirect(url_for("activities"))
    header, rows = override if override else export_rows(acts)
    wb = Workbook()
    ws = wb.active
    ws.title = title[:30]
    ws.sheet_view.rightToLeft = True
    ws.append(header)
    for r in rows:
        ws.append(r)
    for col_cells in ws.columns:
        width = max((len(str(c.value)) for c in col_cells if c.value is not None),
                    default=8)
        ws.column_dimensions[col_cells[0].column_letter].width = min(45, width + 4)
    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return send_file(bio, as_attachment=True, download_name=filename,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# ----------------------------------------------------------------- ورود Excel
@app.route("/import", methods=["GET", "POST"])
@perm_required("can_import")
def import_excel():
    db = get_db()
    domains = db.execute("SELECT * FROM domains WHERE is_active=1 "
                         "ORDER BY sort_order, id").fetchall()
    if request.method == "POST":
        if load_workbook is None:
            flash("کتابخانه openpyxl نصب نیست؛ ورود از Excel ممکن نیست.", "error")
            return redirect(url_for("import_excel"))
        domain_id = request.form.get("domain_id", type=int)
        f = request.files.get("file")
        if not domain_id:
            flash("حوزه را انتخاب کنید.", "error")
            return redirect(url_for("import_excel"))
        if not f or not f.filename:
            flash("فایل Excel یا CSV را انتخاب کنید.", "error")
            return redirect(url_for("import_excel"))
        if not f.filename.lower().endswith((".xlsx", ".xlsm", ".csv")):
            flash("فقط فایل Excel (xlsx) یا CSV پذیرفته می‌شود.", "error")
            return redirect(url_for("import_excel"))
        domain = get_domain_or_404(domain_id)
        data = f.read()
        res = _process_excel(domain, data, f.filename)
        if isinstance(res, dict):
            # ردیف تکراری پیدا شد — پیش از هر ثبتی از کاربر سؤال می‌کنیم
            token = secrets.token_hex(12)
            _ext = os.path.splitext(f.filename)[1].lower() or ".xlsx"
            tmp = os.path.join(UPLOAD_DIR, f"tmp_imp_{g.user['id']}_{token}{_ext}")
            os.makedirs(UPLOAD_DIR, exist_ok=True)
            with open(tmp, "wb") as fh:
                fh.write(data)
            session["imp_tmp"] = {"path": tmp, "domain_id": domain["id"],
                                  "filename": f.filename, "domain_name": domain["name"],
                                  "total": res["total"], "new_n": res["new_n"],
                                  "dups": res["dups"], "errors_n": res["errors_n"]}
            return render_template("import_preview.html", p=session["imp_tmp"],
                                   imp_domains=[{"id": d["id"], "name": d["name"]} for d in domains])
        return redirect(url_for("import_result", import_id=res))
    # پاک‌سازی تنبلی فایل‌های موقت قدیمی (منتظر تأیید ردیف تکراری مانده‌اند)
    try:
        import glob as _glob
        import time as _time
        for pth in _glob.glob(os.path.join(UPLOAD_DIR, "tmp_imp_*")):
            if os.path.getmtime(pth) < _time.time() - 86400:
                os.remove(pth)
    except OSError:
        pass
    return render_template("import_excel.html", domains=domains, icons=DOMAIN_ICONS,
                           imp_domains=[{"id": d["id"], "name": d["name"]} for d in domains])


def _norm_header(s):
    return re.sub(r"\s+", " ", str(s or "").strip())


def _norm_person(s):
    """نرمال‌سازی نام شخص برای تطبیق کارشناس Excel با کاربران سامانه —
    تفاوت فاصله/نیم‌فاصله/حروف عربی (ي ك ة...) و بزرگی‌حروف را نادیده می‌گیرد."""
    s = str(s or "").strip().lower()
    for a, b in (("ي", "ی"), ("ك", "ک"), ("أ", "ا"), ("إ", "ا"), ("آ", "ا"),
                 ("ة", "ه"), ("ؤ", "و"), ("ئ", "ی"), ("۰", "0")):
        s = s.replace(a, b)
    return re.sub(r"[\s‌​‎‏\-_ـ]+", "", s).replace("ي", "ی").replace("ك", "ک")


def _canon_build(db):
    """داده‌های لازم برای یکدست‌سازی نام کارشناس (یک بار در هر درخواست ساخته می‌شود):
    نام‌های کامل = کاربران فعال + نام‌های کارشناسِ ثبت‌شده در داده‌ها؛ نام‌های مستعار کاربران."""
    fulls, user_fns, alias = [], [], {}
    for r in db.execute("SELECT full_name, aliases FROM users WHERE is_active=1"):
        fn = (r["full_name"] or "").strip()
        if fn:
            user_fns.append(fn)
            if fn not in fulls:
                fulls.append(fn)
            for al in re.split(r"[,،;\n]+", r["aliases"] or ""):
                al = al.strip()
                if al:
                    alias[_norm_person(al)] = fn
    for r in db.execute("""SELECT DISTINCT TRIM(v.value) n FROM activity_values v
                           JOIN form_fields ff ON ff.id=v.field_id AND ff.field_key='expert'
                           WHERE TRIM(v.value)<>''"""):
        n = r["n"]
        if n and n not in fulls:
            fulls.append(n)
    return {"user_norms": {_norm_person(f): f for f in user_fns},
            "norms": {f: _norm_person(f) for f in fulls}, "alias": alias, "memo": {}}


_NAME_SPLIT_RE = re.compile(r"[،,;/؛\|\n]|\s+و\s+")


def _name_parts(name):
    """تجزیه سلول کارشناس به بخش‌ها (جداکننده: «،» «،» «/» «؛» «|» خط جدید یا « و »)."""
    return [p.strip() for p in _NAME_SPLIT_RE.split(name) if p.strip()]


def _canon_one(name, c):
    """یکدست‌سازی «یک» نام کارشناس بر اساس داده‌های کش:
    ۱) نظیر دقیقِ نام یک کاربر  ۲) نام مستعار تعریف‌شده برای کاربر
    ۳) پسوندِ یکتای یک نام کاملِ «تک‌نفره» (مثل «رضایی» ← «علی رضایی») — فقط اگر مبهم نباشد
    ۴) واریانت املاییِ دقیقِ یک نام موجود (ی/ک عربی، نیم‌فاصله...)"""
    n = _norm_person(name)
    out = c["user_norms"].get(n)
    if out is None:
        out = c["alias"].get(n)
    if out is None:
        ms = [f for f, nf in c["norms"].items()
              if nf != n and nf.endswith(n) and len(_name_parts(f)) == 1]
        # ابهام بر اساس «مقصد یکتا» سنجیده می‌شود: دو املای یک نام یا نام مستعارِ آن
        # (مثل «علی‌ رضایی» و «ع.رضایی» ← «علی رضایی») مبهم نیستند
        tgts = sorted({c["user_norms"].get(c["norms"][f]) or c["alias"].get(c["norms"][f]) or f
                       for f in ms})
        if len(tgts) == 1:
            out = tgts[0]
    if out is None:
        for f, nf in c["norms"].items():
            if nf == n and f != name:
                out = f
                break
    return out if out is not None else name


def _canon_expert(name):
    """نام نمایشی کارشناس را یکدست می‌کند تا در نمودار/گزارش دو بار نیاید.
    اگر چند نام در یک سلول باشد (با «،» «و» «/» «؛» یا خط جدید)، هر بخش جداگانه
    یکدست می‌شود، جداکننده‌ها یکسان («، ») و بخش‌های تکراری حذف می‌شوند."""
    name = (name or "").strip()
    if not name:
        return name
    c = getattr(g, "_canon_cache", None)
    if c is None:
        c = g._canon_cache = _canon_build(get_db())
    if name in c["memo"]:
        return c["memo"][name]
    parts = _name_parts(name)
    if len(parts) > 1:
        out = "، ".join(dict.fromkeys(_canon_one(p, c) for p in parts))
    else:
        out = _canon_one(name, c)
    c["memo"][name] = out
    return out


def _cell_text(raw):
    """متن تمیز یک سلول اکسل؛ اعداد عددیِ بدون اعشار (مثل 4321.0) صحیح برمی‌گردند."""
    if raw is None:
        return ""
    if isinstance(raw, float) and raw.is_integer():
        raw = int(raw)
    return str(raw).strip()


# نام‌های جایگزین متداول ستون‌ها برای تطبیق هوشمند سرستون اکسل با فیلدهای کلیدی
_EXCEL_KEY_ALIASES = {
    "ticket": ["ticket", "ticket no", "ticket number", "تیکت", "شماره تیکت",
               "شماره تیکت / درخواست", "نام و شماره تیکت"],
    "title": ["title", "subject", "عنوان", "عنوان فعالیت", "عنوان آسیب‌پذیری",
              "محصول", "نام محصول", "product", "product name", "نام آسیب‌پذیری"],
    "date": ["date", "تاریخ", "تاریخ فعالیت", "تاریخ ثبت"],
    "expert": ["expert", "کارشناس", "مسئول", "کارشناس مسئول"],
}


def _map_excel_headers(headers, fields):
    """تطبیق سرستون‌های اکسل با فیلدها: دقیق بر اساس برچسب، سپس نام‌های جایگزین."""
    col_map, name_to_field = {}, {}
    for fld in fields:
        name_to_field.setdefault(_norm_header(fld["label"]), fld)
    status_col = None
    key_fields = {}
    for fld in fields:
        if fld["field_key"] and fld["field_key"] not in key_fields:
            key_fields[fld["field_key"]] = fld
    used = set()
    # پاس ۱: تطابق دقیق برچسب
    for i, h in enumerate(headers):
        if h in name_to_field and name_to_field[h]["id"] not in used:
            col_map[name_to_field[h]["id"]] = i
            used.add(name_to_field[h]["id"])
        elif h == "وضعیت":
            status_col = i
    # پاس ۲: نام‌های جایگزین برای فیلدهای کلیدیِ بدون تطابق
    for i, h in enumerate(headers):
        if i in col_map.values():
            continue
        hl = h.lower()
        for key, aliases in _EXCEL_KEY_ALIASES.items():
            fld = key_fields.get(key)
            if not fld or fld["id"] in used:
                continue
            if hl in aliases or (key == "ticket" and "تیکت" in h):
                col_map[fld["id"]] = i
                used.add(fld["id"])
                break
    return col_map, status_col


def _excel_sig(record, fids_sorted):
    """امضای یک رکورد روی همهٔ فیلدهای فعال حوزه (ستون غایب = رشتهٔ خالی)."""
    return tuple((fid, str(record.get(fid) or "").strip()) for fid in fids_sorted)


def _domain_sigs(db, domain_id, fids_sorted):
    """امضای همهٔ فعالیت‌های موجود حوزه — برای تشخیص ردیف تکراری."""
    m = {}
    for aid, fid, val in db.execute(
            "SELECT a.id, v.field_id, v.value FROM activities a "
            "LEFT JOIN activity_values v ON v.activity_id=a.id WHERE a.domain_id=?",
            (domain_id,)):
        if fid is not None:
            m.setdefault(aid, {})[fid] = val
    return {_excel_sig(vmap, fids_sorted) for vmap in m.values()}


def _domain_sig_map(db, domain_id, fids_sorted):
    """نگاشت امضای فعالیت‌های حوزه به فهرست (id, user_id) —
    برای اصلاح کارشناس ردیف‌های تکراریِ قبلاً ثبت‌شده."""
    m, owners = {}, {}
    for aid, uid, fid, val in db.execute(
            "SELECT a.id, a.user_id, v.field_id, v.value FROM activities a "
            "LEFT JOIN activity_values v ON v.activity_id=a.id WHERE a.domain_id=?",
            (domain_id,)):
        owners[aid] = uid
        if fid is not None:
            m.setdefault(aid, {})[fid] = val
    out = {}
    for aid, vmap in m.items():
        out.setdefault(_excel_sig(vmap, fids_sorted), []).append((aid, owners.get(aid)))
    return out


def _read_table_rows(data, filename):
    """خواندن سطرهای فایل ورودی (xlsx/xlsm/csv) به‌صورت فهرست تاپل؛ در صورت خطا None."""
    if str(filename or "").lower().endswith(".csv"):
        import csv as _csv
        text = None
        for enc in ("utf-8-sig", "utf-8", "cp1256", "windows-1252"):
            try:
                text = data.decode(enc)
                break
            except (UnicodeDecodeError, ValueError):
                continue
        if text is None:
            return None
        sample = text[:4096]
        try:
            dialect = _csv.Sniffer().sniff(sample, delimiters=",;\t|")
        except _csv.Error:
            dialect = _csv.excel
        try:
            rows = [tuple(c.strip() if isinstance(c, str) else c for c in row)
                    for row in _csv.reader(io.StringIO(text), dialect)]
            return [r for r in rows]
        except _csv.Error:
            return None
    try:
        ws = load_workbook(io.BytesIO(data), read_only=True, data_only=True).active
        return list(ws.iter_rows(values_only=True))
    except Exception:
        return None


def _process_excel(domain, data, filename, dup_mode="ask"):
    """پردازش فایل Excel/CSV حوزه.
    dup_mode:
      'ask'   — اگر ردیف تکراری (همهٔ ستون‌ها مشابه یک فعالیت موجود یا تکرار در خود فایل)
                 باشد، چیزی ثبت نمی‌کند و دیکشنری پیش‌نمایش برمی‌گرداند تا از کاربر سؤال شود
      'skip'  — فقط رکوردهای جدید ثبت می‌شوند؛ ردیف‌های تکراری نادیده گرفته می‌شوند
      'force' — همهٔ رکوردهای معتبر ثبت می‌شوند (حتی تکراری‌ها)"""
    db = get_db()
    rows = _read_table_rows(data, filename)
    if rows is None:
        flash("فایل به‌درستی خوانده نشد. یک فایل xlsx یا csv معتبر انتخاب کنید.", "error")
        return db.execute("INSERT INTO excel_imports(domain_id,user_id,filename,total_rows,"
                          "success_rows,error_rows,errors,imported_at) VALUES(?,?,?,?,0,?,?,?)",
                          (domain["id"], g.user["id"], filename, 0, 0,
                           "فایل خوانده نشد", now_iso())).lastrowid

    fields = get_fields(domain["id"])
    while rows and all(c is None or str(c).strip() == "" for c in rows[0]):
        rows.pop(0)
    if not rows:
        flash("فایل خالی است.", "error")
    headers = [_norm_header(c) for c in (rows[0] if rows else [])]
    col_map, status_col = _map_excel_headers(headers, fields)
    if not col_map:
        # فایل ناسازگار: هیچ ردیفی ثبت نمی‌شود؛ فقط لاگ خطا ذخیره می‌شود
        flash("هیچ‌کدام از ستون‌های فایل با فیلدهای این حوزه مطابقت ندارد — "
              "فایل ثبت نشد. از «دانلود قالب نمونه» استفاده کنید.", "error")
        total0 = max(0, len(rows) - 1)
        log_id = db.execute(
            "INSERT INTO excel_imports(domain_id,user_id,filename,total_rows,"
            "success_rows,error_rows,errors,imported_at) VALUES(?,?,?,?,0,?,?,?)",
            (domain["id"], g.user["id"], filename, total0, total0,
             "ستون‌های فایل با فیلدهای حوزه مطابقت ندارد؛ ورود فایل لغو شد.",
             now_iso())).lastrowid
        db.commit()
        return log_id

    errors, success, warns, owner_fix = [], 0, [], 0
    unknown_experts = {}   # نام کارشناسِ تعریف‌نشده → ردیف‌ها (ثبت با همان نامِ فایل)
    total = max(0, len(rows) - 1)
    fids_sorted = sorted(f["id"] for f in fields)
    have_sigs, seen_sigs, seen_raw, sig_map = set(), set(), set(), {}
    if dup_mode != "force":
        sig_map = _domain_sig_map(db, domain["id"], fids_sorted)
        have_sigs = set(sig_map)
    parsed, dups = [], []
    # فیلد «کارشناس» (در صورت وجود در این حوزه) و نگاشت نام → کاربر سامانه
    expert_fid = next((f["id"] for f in fields if (f["field_key"] or "") == "expert"), None)
    if expert_fid is not None and col_map.get(expert_fid) is None and total:
        _elbl = _norm_header(next((f["label"] for f in fields if f["id"] == expert_fid), "کارشناس"))
        warns.append(f"ستون کارشناس (با نام «{_elbl}») در فایل یافت نشد — "
                     f"همهٔ ردیف‌ها به نام شما ثبت شدند. برای ثبت به نام کارشناس هر ردیف، "
                     f"ستونی با همین نام به فایل اضافه کنید.")
    users_by_name = {}
    for r in db.execute("SELECT id, username, full_name, aliases FROM users WHERE is_active=1"):
        users_by_name[_norm_person(r["full_name"])] = r["id"]
        users_by_name.setdefault(_norm_person(r["username"]), r["id"])
        for _al in re.split(r"[,،;\n]+", r["aliases"] or ""):
            if _al.strip():
                users_by_name.setdefault(_norm_person(_al.strip()), r["id"])
    for idx, r in enumerate(rows[1:], start=2):
        if all(c is None or str(c).strip() == "" for c in r):
            total -= 1
            continue
        record, row_err = {}, []
        for fld in fields:
            i = col_map.get(fld["id"])
            raw = r[i] if i is not None and i < len(r) else None
            txt = _cell_text(raw)
            if fld["field_type"] == "date":
                iso = jalali.guess_date_to_g(raw)
                if txt and not iso:
                    row_err.append(f"ستون «{fld['label']}»: تاریخ نامعتبر ({txt})")
                    continue
                if iso:
                    record[fld["id"]] = iso
            else:
                if fld["field_type"] == "number" and txt:
                    txt = jalali.to_ascii_digits(txt)
                if txt:
                    record[fld["id"]] = txt
            # فیلد الزامیِ «غایب از فایل» خطا نیست (ورود تلرانت: ستون‌های موجود خوانده می‌شوند)
            if fld["required"] and col_map.get(fld["id"]) is not None and not record.get(fld["id"]):
                row_err.append(f"ستون «{fld['label']}» خالی یا نامعتبر است")
        if not record and not row_err:
            row_err.append("هیچ مقدار معتبری در ستون‌های این حوزه یافت نشد")
        if not record or (expert_fid is not None and
                          all(fid == expert_fid or not val for fid, val in record.items())):
            # ردیف «روح»: هیچ دادهٔ واقعی (غیر از نام کارشناس) ندارد — ثبت نمی‌شود
            if row_err:
                errors.append(f"ردیف {jalali.fa(idx)}: " + "؛ ".join(row_err))
            continue
        # ردیف‌های دارای نقصِ جزیی «وارد» می‌شوند اما با پرچم «نیازمند اصلاح» (قرمز در لیست)
        flag = 1 if row_err else 0
        if row_err:
            errors.append(f"ردیف {jalali.fa(idx)}: " + "؛ ".join(row_err))
        status = STATUSES[0]
        if status_col is not None and status_col < len(r) and r[status_col]:
            st = _norm_header(r[status_col])
            if st in STATUSES:
                status = st
        if g.user["role"] != "admin" and status == "بررسی شده":
            status = STATUSES[1]
        # نگاشت کارشناس ردیف به کاربر سامانه — پیش‌فرض: خودِ آپلودکننده
        # یکدست‌سازی نام کارشناسِ ثبت‌شده («رضایی» ← «علی رضایی» اگر یکتا باشد)
        # تا از همان ابتدا در آمار و نمودارها دو بار نیاید
        if expert_fid is not None and record.get(expert_fid):
            record[expert_fid] = _canon_expert(record[expert_fid])
        # نکته: نام کارشناسِ داخل فایل به‌هر‌حال در فیلد «کارشناس» ثبت می‌شود و همان
        # در همهٔ نماها (لیست/جزئیات/خروجی/نمودار) نمایش داده می‌شود؛ مالکیت فقط برای
        # سطح دسترسی است و وقتی کاربرِ نظیر در سامانه نیست، به آپلودکننده می‌رسد.
        owner_id, owner_mapped = g.user["id"], False
        if expert_fid is not None:
            ev = str(record.get(expert_fid) or "").strip()
            if ev:
                mu = users_by_name.get(_norm_person(ev))
                if mu is not None:
                    owner_id, owner_mapped = mu, True
                else:
                    unknown_experts.setdefault(ev, []).append(idx)
        # --- تشخیص ردیف تکراری:
        #     • درون فایل: کل سطر خام باید دقیقاً مثل سطر قبلی باشد
        #     • با دیتابیس: مقادیر ستون‌های نگاشت‌شده مشابه یک فعالیت موجود باشد
        sig = _excel_sig(record, fids_sorted)
        raw_sig = "".join(_cell_text(c).strip() for c in r)
        if dup_mode != "force" and (sig in have_sigs or raw_sig in seen_raw):
            seen_raw.add(raw_sig)
            seen_sigs.add(sig)
            dups.append(idx)
            # اصلاح کارشناس فعالیتِ قبلاً ثبت‌شده وقتی فایل نام کارشناس را دارد
            # (به‌ویژه برای داده‌هایی که قبل از این قابلیت وارد شده‌اند)
            if owner_mapped and dup_mode == "skip" and sig in sig_map:
                for aid, ouid in sig_map[sig]:
                    if ouid != owner_id:
                        db.execute("UPDATE activities SET user_id=?, updated_at=? WHERE id=?",
                                   (owner_id, now_iso(), aid))
                        owner_fix += 1
            continue
        seen_raw.add(raw_sig)
        seen_sigs.add(sig)
        parsed.append((record, status, owner_id, flag))
    # کارشناسانِ ناشناخته: یک هشدار تجمیعی به‌ازای هر نام (فعالیت با همان نامِ فایل ثبت می‌شود)
    for _name, _idxs in unknown_experts.items():
        _rt = "، ".join(jalali.fa(i) for i in _idxs[:12]) + ("…" if len(_idxs) > 12 else "")
        warns.append(f"کارشناس «{_name}» ({jalali.fa(len(_idxs))} ردیف: {_rt}) در فهرست "
                     f"کاربران سامانه نیست — فعالیت‌ها با همان نام «{_name}» ثبت شدند "
                     f"و مالکیت مدیریتی‌شان با شماست")
    if dup_mode == "ask" and dups:
        return {"ask": True, "total": total, "new_n": len(parsed),
                "dups": dups, "errors_n": len(errors)}
    flagged_n = sum(1 for _,_,_,fl in parsed if fl)
    for record, status, owner_id, fl in parsed:
        cur = db.execute("INSERT INTO activities(domain_id,user_id,status,created_at,"
                         "updated_at,created_by,flagged) VALUES(?,?,?,?,?,?,?)",
                         (domain["id"], owner_id, status, now_iso(), now_iso(),
                          g.user["id"], fl))
        save_values(cur.lastrowid, record)
        success += 1
    db.commit()
    log_id = db.execute("INSERT INTO excel_imports(domain_id,user_id,filename,total_rows,"
                        "success_rows,error_rows,errors,dup_rows,warns,imported_at) "
                        "VALUES(?,?,?,?,?,?,?,?,?,?)",
                        (domain["id"], g.user["id"], filename, total, success,
                         len(errors), "\n".join(errors[:200]), len(dups),
                         "\n".join(warns[:200]), now_iso())).lastrowid
    db.commit()
    if total == 0:
        flash("فایلی برای پردازش یافت نشد.", "error")
    elif errors:
        flash(f"{jalali.fa(success)} فعالیت ثبت شد، {jalali.fa(len(errors))} ردیف دارای "
              f"خطا بودند.", "warning")
    else:
        flash(f"{jalali.fa(success)} فعالیت با موفقیت ثبت شد.", "success")
    if dups:
        flash(f"{jalali.fa(len(dups))} ردیف تکراری (همهٔ ستون‌هایشان مشابه یک فعالیت "
              f"موجود بود) نادیده گرفته شد.", "warning")
    if owner_fix:
        flash(f"کارشناس {jalali.fa(owner_fix)} فعالیتِ قبلاً ثبت‌شده مطابق ستون کارشناس "
              f"فایل اصلاح شد.", "success")
    if flagged_n:
        flash(f"{jalali.fa(flagged_n)} ردیف دارای نقص وارد شد و با برچسب قرمز "
              f"«نیازمند اصلاح» علامت خورد — بعداً آن‌ها را ویرایش و اصلاح کنید.", "warning")
    if unknown_experts:
        _n = sum(len(v) for v in unknown_experts.values())
        valid_names = "، ".join(sorted({r["full_name"] for r in db.execute(
            "SELECT full_name FROM users WHERE is_active=1")}))
        flash(f"{jalali.fa(_n)} ردیف کارشناس‌شان در فهرست کاربران سامانه نبود؛ فعالیت‌ها "
              f"با همان نامِ داخل فایل ثبت شدند و فقط مالکیت مدیریتی آن‌ها با شماست. "
              f"برای اتصال به حساب کاربری، نامِ داخل فایل باید دقیقاً برابر «نام کامل» یا "
              f"«نام کاربری» یکی از این کاربران باشد: {valid_names}", "warning")
    return log_id


@app.route("/import/confirm", methods=["POST"])
@perm_required("can_import")
def import_confirm():
    """تأیید نهایی ورود فایل پس از دیدن ردیف‌های تکراری."""
    info = session.pop("imp_tmp", None)
    choice = request.form.get("choice", "cancel")

    def _cleanup():
        try:
            path = (info or {}).get("path", "")
            if path and os.path.exists(path):
                os.remove(path)
        except OSError:
            pass

    if not info:
        flash("جلسهٔ ورود فایل منقضی شده است؛ دوباره فایل را انتخاب کنید.", "error")
        return redirect(url_for("import_excel"))
    if choice == "cancel":
        _cleanup()
        flash("ورود فایل لغو شد؛ هیچ ردیفی ثبت نشد.", "warning")
        return redirect(url_for("import_excel"))
    path = info.get("path", "")
    try:
        import time as _time
        fresh = os.path.exists(path) and os.path.getmtime(path) > _time.time() - 7200
    except OSError:
        fresh = False
    if not fresh:
        _cleanup()
        flash("مهلتٔ تأیید فایل تمام شده است؛ دوباره فایل را انتخاب کنید.", "error")
        return redirect(url_for("import_excel"))
    domain = get_db().execute("SELECT * FROM domains WHERE id=?",
                              (info["domain_id"],)).fetchone()
    if not domain:
        _cleanup()
        flash("حوزهٔ انتخابی دیگر وجود ندارد.", "error")
        return redirect(url_for("import_excel"))
    with open(path, "rb") as fh:
        data = fh.read()
    mode = "force" if choice == "force" else "skip"
    result_id = _process_excel(domain, data, info["filename"], dup_mode=mode)
    _cleanup()
    return redirect(url_for("import_result", import_id=result_id))


@app.route("/import/<int:import_id>")
@login_required
def import_result(import_id):
    db = get_db()
    imp = db.execute("""SELECT e.*, d.name domain_name, u.full_name FROM excel_imports e
                        LEFT JOIN domains d ON d.id=e.domain_id
                        LEFT JOIN users u ON u.id=e.user_id WHERE e.id=?""",
                     (import_id,)).fetchone()
    if not imp:
        abort(404)
    if g.user["role"] != "admin" and imp["user_id"] != g.user["id"]:
        abort(403)
    errs = [e for e in (imp["errors"] or "").split("\n") if e]
    warns = [w for w in (imp["warns"] or "").split("\n") if w]
    return render_template("import_result.html", imp=imp, errs=errs, warns=warns)


@app.route("/import/template/<int:domain_id>")
@login_required
def import_template(domain_id):
    if Workbook is None:
        abort(404)
    domain = get_domain_or_404(domain_id)
    fields = get_fields(domain_id)
    wb = Workbook()
    ws = wb.active
    ws.title = domain["name"][:30]
    ws.sheet_view.rightToLeft = True
    hdrs = [f["label"] for f in fields]
    if "وضعیت" not in hdrs:
        hdrs = hdrs + ["وضعیت"]
    ws.append(hdrs)
    sample = []
    for f in fields:
        if f["field_type"] == "date":
            y, m, d = jalali.today_jalali()
            sample.append(f"{y:04d}/{m:02d}/{d:02d}")
        elif f["field_type"] == "select":
            opts = field_options(f)
            sample.append(opts[0] if opts else "")
        elif f["field_type"] == "number":
            sample.append(3)
        else:
            sample.append("نمونه")
    ws.append(sample + ([] if "وضعیت" in [f["label"] for f in fields] else [STATUSES[0]]))
    for col_cells in ws.columns:
        ws.column_dimensions[col_cells[0].column_letter].width = 18
    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return send_file(bio, as_attachment=True,
                     download_name=f"قالب_{domain['name']}.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# --------------------------------------------------------------------- گزارش‌ها
@app.route("/reports")
@login_required
def reports():
    where, params = build_filters()
    db = get_db()
    total = db.execute(f"""SELECT COUNT(*) c FROM activities a
                           JOIN users u ON u.id=a.user_id WHERE {where}""",
                       params).fetchone()["c"]
    acts = query_activities(where, params)
    export = request.args.get("export")
    cols = export_cols_param()
    if export == "excel":
        h, r = export_rows(acts)
        if cols:
            h, r = select_cols(h, r, cols)
        return export_excel(acts, "report.xlsx", "گزارش فعالیت‌ها",
                            override=(h, r))
    if export == "csv":
        header, rows = export_rows(acts)
        if cols:
            header, rows = select_cols(header, rows, cols)
        out = io.StringIO()
        w = csv.writer(out)
        w.writerow(header)
        w.writerows(rows)
        return ("\ufeff" + out.getvalue(), 200,
                {"Content-Type": "text/csv; charset=utf-8",
                 "Content-Disposition": "attachment; filename=report.csv"})
    if export == "pdf":
        header, rows = export_rows(acts)
        if cols:
            header, rows = select_cols(header, rows, cols)
        return render_template("report_print.html", header=header, rows=rows,
                               title="گزارش فعالیت‌ها", charts=build_report_charts(acts),
                               filter_text=filters_summary(),
                               sys_name=get_setting("system_name"), autoprint=True)
    domains = db.execute("SELECT * FROM domains WHERE is_active=1 "
                         "ORDER BY sort_order, id").fetchall()
    users = db.execute("SELECT id,full_name FROM users WHERE is_active=1 "
                       "ORDER BY full_name").fetchall()
    export_headers = export_rows(acts)[0] if acts else []
    # داده JSON برای جدول تعاملی Vue 3 (افزودنی)
    acts_json = [{
        "id": a["id"],
        "title": a["title"] or "بدون عنوان",
        "domain": a["domain_name"],
        "icon": domain_icon(a["domain_name"]),
        "expert": _canon_expert((a["expert_txt"] or a["expert_name"]) if "expert_txt" in a.keys() else a["expert_name"]),
        "status": a["status"],
        "date": _jdate(a["date"]) if a["date"] else "—",
        "date_key": a["date"] or "",
        "ticket": a["ticket"] or "",
        "atts": a["att_count"] or 0,
        "view": url_for("activity_view", activity_id=a["id"]),
    } for a in acts]
    orgs = db.execute("SELECT * FROM orgs ORDER BY sort_order, id").fetchall()
    return render_template("reports.html", acts=acts, total=total, domains=domains,
                           users=users, export_headers=export_headers,
                           acts_json=acts_json, orgs=orgs)


# --------------------------------------------------------------------- کاربران
# --------------------------------------------------------------------- پروفایل
@app.route("/profile", methods=["GET", "POST"])
@login_required
def profile():
    db = get_db()
    if request.method == "POST":
        action = request.form.get("action", "info")
        if action == "info":
            full_name = request.form.get("full_name", "").strip()
            if len(full_name) < 2:
                flash("نام نمایشی معتبر نیست.", "error")
            else:
                db.execute("UPDATE users SET full_name=? WHERE id=?",
                           (full_name, g.user["id"]))
                db.commit()
                flash("مشخصات شما به‌روزرسانی شد.", "success")
        elif action == "password":
            cur = request.form.get("current_password", "")
            new = request.form.get("new_password", "")
            rep = request.form.get("repeat_password", "")
            if not check_password_hash(g.user["password_hash"], cur):
                flash("رمز عبور فعلی اشتباه است.", "error")
            elif len(new) < 4:
                flash("رمز عبور جدید باید دست‌کم ۴ نویسه باشد.", "error")
            elif new != rep:
                flash("تکرار رمز عبور جدید با رمز جدید مطابقت ندارد.", "error")
            else:
                db.execute("UPDATE users SET password_hash=? WHERE id=?",
                           (generate_password_hash(new), g.user["id"]))
                db.commit()
                flash("رمز عبور شما تغییر کرد.", "success")
        return redirect(url_for("profile"))
    sup = None
    if g.user["is_trainee"] and g.user["supervisor_id"]:
        sup = db.execute("SELECT full_name, username FROM users WHERE id=?",
                         (g.user["supervisor_id"],)).fetchone()
    return render_template("profile.html", sup=sup)


@app.route("/manage")
@admin_required
def manage():
    """پیشخوان مدیریت — هاب دسترسی سریع به بخش‌های مدیریتی (فقط نمایش)."""
    db = get_db()
    stats = {
        "users": db.execute("SELECT COUNT(*) c FROM users WHERE is_active=1").fetchone()["c"],
        "domains": db.execute("SELECT COUNT(*) c FROM domains WHERE is_active=1").fetchone()["c"],
        "fields": db.execute("SELECT COUNT(*) c FROM form_fields").fetchone()["c"],
        "imports": db.execute("SELECT COUNT(*) c FROM excel_imports").fetchone()["c"],
        "trainees": db.execute("SELECT COUNT(*) c FROM users WHERE is_active=1 AND is_trainee=1").fetchone()["c"],
    }
    teams = [dict(r) for r in db.execute("""
        SELECT u.id, u.full_name, u.username, u.role,
               (SELECT COUNT(*) FROM users t WHERE t.supervisor_id=u.id AND t.is_active=1) tc
        FROM users u WHERE u.is_trainee=0 AND u.is_active=1
        ORDER BY tc DESC, u.id""").fetchall()]
    names = {}
    for t in db.execute("SELECT full_name, supervisor_id FROM users "
                        "WHERE is_trainee=1 AND is_active=1 ORDER BY id").fetchall():
        names.setdefault(t["supervisor_id"], []).append(t["full_name"])
    for tm in teams:
        tm["names"] = names.get(tm["id"], [])
        tm["edit"] = url_for("user_edit", user_id=tm["id"])
    return render_template("manage.html", stats=stats, teams=teams)


@app.route("/users")
@admin_required
def users():
    rows = get_db().execute("""SELECT u.*, (SELECT COUNT(*) FROM activities a
            WHERE a.user_id=u.id) act_count,
            (SELECT COUNT(*) FROM users t WHERE t.supervisor_id=u.id AND t.is_active=1) trainee_count,
            (SELECT s.full_name FROM users s WHERE s.id=u.supervisor_id) supervisor_name
            FROM users u ORDER BY u.id""").fetchall()
    users_json = [{
        "id": u["id"], "username": u["username"], "full_name": u["full_name"],
        "role": u["role"], "act_count": u["act_count"],
        "is_active": bool(u["is_active"]), "self": u["id"] == g.user["id"],
        "is_trainee": bool(u["is_trainee"]),
        "trainee_count": u["trainee_count"],
        "supervisor_name": u["supervisor_name"] or "",
        "can_add": bool(u["can_add"]), "can_edit": bool(u["can_edit"]),
        "can_delete": bool(u["can_delete"]), "can_import": bool(u["can_import"]),
        "edit": url_for("user_edit", user_id=u["id"]),
        "toggle": url_for("user_toggle", user_id=u["id"]),
        "delete": url_for("user_delete", user_id=u["id"]),
    } for u in rows]
    return render_template("users.html", users=rows, users_json=users_json)


def _valid_supervisor(db, sup_id, exclude_id):
    """شناسه سرپرست معتبر: کاربر فعال غیرکارآموز، غیر از خود کاربر."""
    if not sup_id:
        return None
    s = db.execute("SELECT id FROM users WHERE id=? AND is_trainee=0 AND id!=?",
                   (sup_id, exclude_id)).fetchone()
    return s["id"] if s else None


def _sup_candidates(db, exclude_id):
    return db.execute("SELECT id, full_name, username FROM users "
                      "WHERE is_trainee=0 AND id!=? ORDER BY id",
                      (exclude_id,)).fetchall()


@app.route("/users/new", methods=["GET", "POST"])
@admin_required
def user_new():
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        full_name = request.form.get("full_name", "").strip()
        aliases = request.form.get("aliases", "").strip()
        role = request.form.get("role", "expert")
        password = request.form.get("password", "")
        is_trainee = 1 if request.form.get("is_trainee") else 0
        if request.form.get("perm_form"):  # فرم جدید: وضعیت چک‌باکس‌ها معتبر است
            perms = {c: (1 if request.form.get(c) else 0) for c in PERM_COLS}
        else:  # پست‌های قدیمی/برنامه‌ای: پیش‌فرض رفتار قبلی حفظ شود
            perms = {"can_add": 1, "can_edit": 1, "can_delete": 0, "can_import": 1}
        db = get_db()
        sup_id = _valid_supervisor(db, request.form.get("supervisor_id", type=int), -1) \
            if is_trainee else None
        if not username or not full_name or not password:
            flash("نام کاربری، نام و رمز عبور الزامی است.", "error")
        elif role not in ("admin", "expert"):
            flash("نقش نامعتبر است.", "error")
        elif db.execute("SELECT 1 FROM users WHERE username=?", (username,)).fetchone():
            flash("این نام کاربری قبلاً ثبت شده است.", "error")
        elif is_trainee and request.form.get("supervisor_id") and not sup_id:
            flash("سرپرست انتخاب‌شده معتبر نیست.", "error")
        else:
            db.execute("INSERT INTO users(username,password_hash,full_name,role,"
                       "is_trainee,supervisor_id,can_add,can_edit,can_delete,can_import,"
                       "aliases,created_at) VALUES(?,?,?,?,?,?,?,?,?,?,?,?)",
                       (username, generate_password_hash(password), full_name, role,
                        is_trainee, sup_id, perms["can_add"], perms["can_edit"],
                        perms["can_delete"], perms["can_import"], aliases, now_iso()))
            db.commit()
            flash("کاربر ایجاد شد.", "success")
            return redirect(url_for("users"))
    sups = _sup_candidates(get_db(), -1)
    return render_template("user_form.html", u=None, sups=sups,
                           sups_json=[{"id": x["id"], "full_name": x["full_name"],
                                       "username": x["username"]} for x in sups])


@app.route("/users/<int:user_id>/edit", methods=["GET", "POST"])
@admin_required
def user_edit(user_id):
    db = get_db()
    u = db.execute("SELECT * FROM users WHERE id=?", (user_id,)).fetchone()
    if not u:
        abort(404)
    if request.method == "POST":
        full_name = request.form.get("full_name", "").strip()
        aliases = request.form.get("aliases", u["aliases"] or "").strip()
        role = request.form.get("role", u["role"])
        password = request.form.get("password", "")
        is_self = user_id == g.user["id"]
        if is_self and role != "admin":
            role = "admin"  # مدیر نمی‌تواند خودش را تنزل دهد
        is_trainee = 1 if request.form.get("is_trainee") else 0
        sup_id = _valid_supervisor(db, request.form.get("supervisor_id", type=int), user_id) \
            if is_trainee else None
        perms = {c: (1 if request.form.get(c) else 0) for c in PERM_COLS}
        if not full_name:
            flash("نام کامل الزامی است.", "error")
        elif is_trainee and request.form.get("supervisor_id") and not sup_id:
            flash("سرپرست انتخاب‌شده معتبر نیست.", "error")
        else:
            db.execute("UPDATE users SET full_name=?, role=?, aliases=? WHERE id=?",
                       (full_name, role, aliases, user_id))
            if not is_self and request.form.get("perm_form"):
                # نوع حساب، سرپرست و مجوزها (حساب خود مدیر و پست‌های قدیمی مستثنا)
                db.execute("UPDATE users SET is_trainee=?, supervisor_id=?, can_add=?,"
                           " can_edit=?, can_delete=?, can_import=? WHERE id=?",
                           (is_trainee, sup_id, perms["can_add"], perms["can_edit"],
                            perms["can_delete"], perms["can_import"], user_id))
                if is_trainee:
                    db.execute("UPDATE users SET supervisor_id=NULL WHERE supervisor_id=?",
                               (user_id,))
            if password:
                db.execute("UPDATE users SET password_hash=? WHERE id=?",
                           (generate_password_hash(password), user_id))
            db.commit()
            flash("کاربر به‌روزرسانی شد.", "success")
            return redirect(url_for("users"))
    sups = _sup_candidates(db, user_id)
    return render_template("user_form.html", u=u, sups=sups,
                           sups_json=[{"id": x["id"], "full_name": x["full_name"],
                                       "username": x["username"]} for x in sups])


@app.route("/users/<int:user_id>/toggle", methods=["POST"])
@admin_required
def user_toggle(user_id):
    if user_id == g.user["id"]:
        flash("نمی‌توانید حساب خودتان را غیرفعال کنید.", "error")
        return redirect(url_for("users"))
    db = get_db()
    db.execute("UPDATE users SET is_active = 1 - is_active WHERE id=?", (user_id,))
    db.commit()
    return redirect(url_for("users"))


@app.route("/users/<int:user_id>/delete", methods=["POST"])
@admin_required
def user_delete(user_id):
    if user_id == g.user["id"]:
        flash("نمی‌توانید حساب خودتان را حذف کنید.", "error")
        return redirect(url_for("users"))
    db = get_db()
    c = db.execute("SELECT COUNT(*) c FROM activities WHERE user_id=?", (user_id,)
                   ).fetchone()["c"]
    if c:
        flash("این کاربر فعالیت ثبت‌شده دارد؛ به‌جای حذف، آن را غیرفعال کنید.", "error")
        return redirect(url_for("users"))
    db.execute("UPDATE users SET supervisor_id=NULL WHERE supervisor_id=?", (user_id,))
    db.execute("DELETE FROM users WHERE id=?", (user_id,))
    db.commit()
    flash("کاربر حذف شد.", "success")
    return redirect(url_for("users"))


# -------------------------------------------------------------------- تنظیمات
@app.route("/settings", methods=["GET", "POST"])
@admin_required
def settings():
    if request.method == "POST":
        name = request.form.get("system_name", "").strip()
        if name:
            set_setting("system_name", name)
        mb = request.form.get("max_upload_mb", "10").strip()
        if jalali.to_ascii_digits(mb).isdigit() and int(jalali.to_ascii_digits(mb)) > 0:
            set_setting("max_upload_mb", str(int(jalali.to_ascii_digits(mb))))
        fmts = request.form.get("allowed_formats", "").lower()
        fmts = ",".join(x.strip().lstrip(".") for x in re.split(r"[,،;\s]+", fmts) if x.strip())
        if fmts:
            set_setting("allowed_formats", fmts)
        logo = request.files.get("logo")
        if logo and logo.filename:
            ext = logo.filename.rsplit(".", 1)[-1].lower() if "." in logo.filename else ""
            if ext in ("png", "jpg", "jpeg", "gif", "webp", "svg"):
                for old in os.listdir(UPLOAD_DIR):
                    if old.startswith("logo."):
                        os.remove(os.path.join(UPLOAD_DIR, old))
                logo.save(os.path.join(UPLOAD_DIR, f"logo.{ext}"))
                set_setting("logo_ext", ext)
            else:
                flash("فرمت لوگو باید تصویری باشد.", "error")
        flash("تنظیمات ذخیره شد.", "success")
        return redirect(url_for("settings"))
    return render_template("settings.html",
                           sys_name=get_setting("system_name"),
                           max_mb=max_upload_mb(), formats=get_setting("allowed_formats"),
                           has_logo=os.path.exists(
                               os.path.join(UPLOAD_DIR, f"logo.{get_setting('logo_ext')}")))


@app.route("/logo")
def logo():
    path = os.path.join(UPLOAD_DIR, f"logo.{get_setting('logo_ext')}")
    if os.path.exists(path):
        return send_file(path)
    abort(404)


@app.route("/domains", methods=["GET", "POST"])
@admin_required
def domains_page():
    db = get_db()
    if request.method == "POST":
        name = request.form.get("name", "").strip()
        if not name:
            flash("نام حوزه الزامی است.", "error")
        elif db.execute("SELECT 1 FROM domains WHERE name=?", (name,)).fetchone():
            flash("حوزه‌ای با این نام وجود دارد.", "error")
        else:
            org_id = request.form.get("org_id", type=int)
            if not org_id or not db.execute("SELECT 1 FROM orgs WHERE id=?", (org_id,)).fetchone():
                org_id = db.execute("SELECT id FROM orgs ORDER BY sort_order, id LIMIT 1").fetchone()["id"]
            mx = db.execute("SELECT COALESCE(MAX(sort_order),0) m FROM domains").fetchone()["m"]
            db.execute("INSERT INTO domains(name,sort_order,org_id) VALUES(?,?,?)",
                       (name, mx + 1, org_id))
            db.commit()
            flash("حوزه جدید افزوده شد. از «مدیریت فیلدها» فیلدهای آن را تعریف کنید.", "success")
        return redirect(url_for("domains_page"))
    rows = db.execute("""SELECT d.*, o.name org_name,
        (SELECT COUNT(*) FROM activities a WHERE a.domain_id=d.id) act_count,
        (SELECT COUNT(*) FROM form_fields f WHERE f.domain_id=d.id AND f.is_active=1) field_count
        FROM domains d LEFT JOIN orgs o ON o.id=d.org_id
        ORDER BY COALESCE(o.sort_order,99), d.sort_order, d.id""").fetchall()
    orgs = db.execute("SELECT * FROM orgs ORDER BY sort_order, id").fetchall()
    org_counts = {}
    for _d in rows:
        org_counts[_d["org_id"]] = org_counts.get(_d["org_id"], 0) + 1
    # داده JSON برای جدول تعاملی Vue 3 (افزودنی)
    domains_json = [{
        "id": d["id"], "name": d["name"], "icon": domain_icon(d["name"]),
        "org": d["org_name"] or "—", "org_id": d["org_id"] or 0,
        "field_count": d["field_count"], "act_count": d["act_count"],
        "is_active": bool(d["is_active"]),
        "edit": url_for("domain_edit", domain_id=d["id"]),
        "toggle": url_for("domain_toggle", domain_id=d["id"]),
        "delete": url_for("domain_delete", domain_id=d["id"]),
        "fields": url_for("fields_page", domain_id=d["id"]),
    } for d in rows]
    orgs_json = [{"id": o["id"], "name": o["name"]} for o in orgs]
    return render_template("domains.html", domains=rows, icons=DOMAIN_ICONS,
                           domains_json=domains_json, orgs=orgs, orgs_json=orgs_json,
                           org_counts=org_counts)


@app.route("/domains/<int:domain_id>/edit", methods=["POST"])
@admin_required
def domain_edit(domain_id):
    d = get_domain_or_404(domain_id)
    name = request.form.get("name", "").strip()
    db = get_db()
    if name and name != d["name"] and db.execute(
            "SELECT 1 FROM domains WHERE name=? AND id<>?", (name, domain_id)).fetchone():
        flash("حوزه‌ای با این نام وجود دارد.", "error")
    elif name:
        db.execute("UPDATE domains SET name=? WHERE id=?", (name, domain_id))
        db.commit()
        flash("حوزه ویرایش شد.", "success")
    org_id = request.form.get("org_id", type=int)
    if org_id and org_id != d["org_id"] and db.execute(
            "SELECT 1 FROM orgs WHERE id=?", (org_id,)).fetchone():
        db.execute("UPDATE domains SET org_id=? WHERE id=?", (org_id, domain_id))
        db.commit()
        flash("مرکز حوزه تغییر کرد.", "success")
    return redirect(request.referrer or url_for("domains_page"))


@app.route("/orgs/add", methods=["POST"])
@admin_required
def org_add():
    name = request.form.get("name", "").strip()
    db = get_db()
    if not name:
        flash("نام مرکز الزامی است.", "error")
    elif db.execute("SELECT 1 FROM orgs WHERE name=?", (name,)).fetchone():
        flash("مرکزی با این نام وجود دارد.", "error")
    else:
        mx = db.execute("SELECT COALESCE(MAX(sort_order),0) m FROM orgs").fetchone()["m"]
        db.execute("INSERT INTO orgs(name,sort_order) VALUES(?,?)", (name, mx + 1))
        db.commit()
        flash("مرکز جدید افزوده شد — حالا می‌توانید برایش حوزه بسازید.", "success")
    return redirect(url_for("domains_page"))


@app.route("/orgs/<int:org_id>/delete", methods=["POST"])
@admin_required
def org_delete(org_id):
    db = get_db()
    o = db.execute("SELECT * FROM orgs WHERE id=?", (org_id,)).fetchone()
    if not o:
        abort(404)
    if db.execute("SELECT 1 FROM domains WHERE org_id=? LIMIT 1", (org_id,)).fetchone():
        flash("این مرکز حوزه دارد و قابل حذف نیست — ابتدا حوزه‌هایش را به مرکز دیگر منتقل کنید.",
              "error")
    else:
        db.execute("DELETE FROM orgs WHERE id=?", (org_id,))
        db.commit()
        flash(f"مرکز «{o['name']}» حذف شد.", "success")
    return redirect(url_for("domains_page"))


@app.route("/domains/<int:domain_id>/toggle", methods=["POST"])
@admin_required
def domain_toggle(domain_id):
    get_domain_or_404(domain_id)
    db = get_db()
    db.execute("UPDATE domains SET is_active = 1 - is_active WHERE id=?", (domain_id,))
    db.commit()
    return redirect(request.referrer or url_for("domains_page"))


@app.route("/domains/<int:domain_id>/delete", methods=["POST"])
@admin_required
def domain_delete(domain_id):
    db = get_db()
    get_domain_or_404(domain_id)
    c = db.execute("SELECT COUNT(*) c FROM activities WHERE domain_id=?",
                   (domain_id,)).fetchone()["c"]
    if c:
        flash("این حوزه دارای فعالیت ثبت‌شده است؛ به‌جای حذف، آن را غیرفعال کنید.", "error")
        return redirect(url_for("domains_page"))
    db.execute("DELETE FROM form_fields WHERE domain_id=?", (domain_id,))
    db.execute("DELETE FROM domains WHERE id=?", (domain_id,))
    db.commit()
    flash("حوزه حذف شد.", "success")
    return redirect(url_for("domains_page"))


@app.route("/domains/<int:domain_id>/fields")
@admin_required
def fields_page(domain_id):
    domain = get_domain_or_404(domain_id)
    rows = get_db().execute("""SELECT f.*, (SELECT COUNT(*) FROM activity_values v
        WHERE v.field_id=f.id) value_count FROM form_fields f WHERE f.domain_id=?
        ORDER BY f.sort_order, f.id""", (domain_id,)).fetchall()
    # داده JSON برای فهرست تعاملی Vue 3 (افزودنی)
    type_fa = {"text": "متن", "textarea": "متن بلند", "number": "عدد", "file": "فایل",
               "date": "تاریخ", "select": "لیست کشویی"}
    fields_vjson = [{
        "id": f["id"], "label": f["label"], "type": f["field_type"],
        "type_fa": type_fa.get(f["field_type"], f["field_type"]),
        "section": f["section"] or "", "required": bool(f["required"]),
        "is_active": bool(f["is_active"]), "value_count": f["value_count"],
        "options": field_options(f),
        "move": url_for("field_move", field_id=f["id"]),
        "edit": url_for("field_edit", field_id=f["id"]),
        "delete": url_for("field_delete", field_id=f["id"]),
    } for f in rows]
    return render_template("fields.html", domain=domain, fields=rows,
                           icons=DOMAIN_ICONS, fields_vjson=fields_vjson,
                           sec_choices=[SEC_REQUEST, SEC_DELIVERY])


@app.route("/domains/<int:domain_id>/fields/add", methods=["POST"])
@admin_required
def field_add(domain_id):
    get_domain_or_404(domain_id)
    label = request.form.get("label", "").strip()
    ftype = request.form.get("field_type", "text")
    options = [o.strip() for o in request.form.get("options", "").split("\n") if o.strip()]
    section = request.form.get("section", "").strip()
    if ftype not in ("text", "textarea", "number", "date", "select", "file"):
        ftype = "text"
    if not label:
        flash("نام فیلد الزامی است.", "error")
        return redirect(url_for("fields_page", domain_id=domain_id))
    db = get_db()
    mx = db.execute("SELECT COALESCE(MAX(sort_order),0) m FROM form_fields "
                    "WHERE domain_id=?", (domain_id,)).fetchone()["m"]
    db.execute("INSERT INTO form_fields(domain_id,label,field_type,section,options,required,"
               "sort_order) VALUES(?,?,?,?,?,?,?)",
               (domain_id, label, ftype, section, json.dumps(options, ensure_ascii=False),
                1 if request.form.get("required") else 0, mx + 1))
    db.commit()
    flash("فیلد افزوده شد.", "success")
    return redirect(request.referrer or url_for("fields_page", domain_id=domain_id))


@app.route("/fields/<int:field_id>/edit", methods=["POST"])
@admin_required
def field_edit(field_id):
    db = get_db()
    f = db.execute("SELECT * FROM form_fields WHERE id=?", (field_id,)).fetchone()
    if not f:
        abort(404)
    label = request.form.get("label", "").strip()
    options = [o.strip() for o in request.form.get("options", "").split("\n") if o.strip()]
    section = request.form.get("section", f["section"] or "").strip()
    if label:
        db.execute("UPDATE form_fields SET label=?, section=?, options=?, required=?, is_active=? "
                   "WHERE id=?",
                   (label, section, json.dumps(options, ensure_ascii=False),
                    1 if request.form.get("required") else 0,
                    1 if request.form.get("is_active") else 0, field_id))
        db.commit()
        flash("فیلد ویرایش شد.", "success")
    return redirect(request.referrer or url_for("fields_page", domain_id=f["domain_id"]))


@app.route("/fields/<int:field_id>/move", methods=["POST"])
@admin_required
def field_move(field_id):
    direction = request.form.get("direction")
    db = get_db()
    f = db.execute("SELECT * FROM form_fields WHERE id=?", (field_id,)).fetchone()
    if not f:
        abort(404)
    sib = db.execute("""SELECT * FROM form_fields WHERE domain_id=? AND
        sort_order {} ? ORDER BY sort_order {} LIMIT 1""".format(
        "<" if direction == "up" else ">", "DESC" if direction == "up" else "ASC"),
        (f["domain_id"], f["sort_order"])).fetchone()
    if sib:
        db.execute("UPDATE form_fields SET sort_order=? WHERE id=?",
                   (sib["sort_order"], f["id"]))
        db.execute("UPDATE form_fields SET sort_order=? WHERE id=?",
                   (f["sort_order"], sib["id"]))
        db.commit()
    return redirect(request.referrer or url_for("fields_page", domain_id=f["domain_id"]))


@app.route("/fields/<int:field_id>/delete", methods=["POST"])
@admin_required
def field_delete(field_id):
    db = get_db()
    f = db.execute("SELECT * FROM form_fields WHERE id=?", (field_id,)).fetchone()
    if not f:
        abort(404)
    c = db.execute("SELECT COUNT(*) c FROM activity_values WHERE field_id=?",
                   (field_id,)).fetchone()["c"]
    if c:
        flash("این فیلد در فعالیت‌های ثبت‌شده مقدار دارد. برای حفظ داده‌ها، "
              "به‌جای حذف آن را غیرفعال کنید.", "error")
        return redirect(url_for("fields_page", domain_id=f["domain_id"]))
    db.execute("DELETE FROM form_fields WHERE id=?", (field_id,))
    db.commit()
    flash("فیلد حذف شد.", "success")
    return redirect(request.referrer or url_for("fields_page", domain_id=f["domain_id"]))


# ------------------------------------------------------------------ قالب‌ها
@app.template_filter("jdate")
def _jdate(iso):
    return jalali.fa(jalali.g_str_to_j(iso))


@app.template_filter("jdatetime")
def _jdatetime(iso):
    if not iso:
        return "—"
    date_part = jalali.fa(jalali.g_str_to_j(iso))
    m = re.search(r" (\d{2}:\d{2})", str(iso))
    return f"{date_part} {jalali.fa(m.group(1))}" if m else date_part


@app.template_filter("fa")
def _fa(v):
    return jalali.fa(v)


@app.template_filter("tojson_attr")
def _tojson_sa(v):
    return json.dumps(v, ensure_ascii=False)


@app.template_filter("tojson_script")
def _tojson_script(v):
    """JSON امن برای تگ <script type="application/json"> — حروف فارسی خوانا می‌مانند."""
    from markupsafe import Markup
    return Markup(json.dumps(v, ensure_ascii=False).replace("</", "<\\/"))


@app.context_processor
def inject_globals():
    # شمارنده تسک‌های بازِ تخصیص‌یافته به کاربر جاری (برای نشان منو) — فقط افزودنی
    nav_tasks = 0
    try:
        u = getattr(g, "user", None)
        if u:
            nav_tasks = get_db().execute(
                """SELECT COUNT(*) c FROM activities
                   WHERE user_id=? AND created_by IS NOT NULL AND created_by != user_id
                     AND status != 'انجام شده'""", [u["id"]]).fetchone()["c"]
    except Exception:
        nav_tasks = 0
    user = getattr(g, "user", None) or (current_user() if request else None)
    perms = {"add": True, "edit": True, "delete": True, "import": True}
    if user and user["role"] != "admin":
        perms = {"add": bool(user["can_add"]), "edit": bool(user["can_edit"]),
                 "delete": bool(user["can_delete"]), "import": bool(user["can_import"])}
    return {"sys_name": get_setting("system_name"), "STATUSES": STATUSES,
            "perms": perms,
            "domain_icon": domain_icon, "has_logo": os.path.exists(
                os.path.join(UPLOAD_DIR, f"logo.{get_setting('logo_ext')}")),
            "asset_v": ASSET_V, "jalali_months": jalali.MONTH_NAMES,
            "SEC_REQUEST_NAME": SEC_REQUEST, "SEC_DELIVERY_NAME": SEC_DELIVERY,
            "today_jalali": jalali.today_jalali(),
            "today_fa_str": jalali.fa("%04d/%02d/%02d" % jalali.today_jalali()),
            "field_options": field_options, "nav_tasks": nav_tasks,
            "base_cols": BASE_COLS}


@app.after_request
def _no_store_dynamic(resp):
    """صفحات HTML پویا هرگز کش نشوند — با Back/Forward و تب‌های باز، همیشه داده تازه."""
    if resp.headers.get("Content-Type", "").startswith("text/html"):
        resp.headers["Cache-Control"] = "no-store, no-cache, must-revalidate"
        resp.headers["Pragma"] = "no-cache"
        resp.headers["Expires"] = "0"
    return resp


@app.errorhandler(403)
def _403(_e):
    if not getattr(g, "user", None):
        g.user = current_user()
    if not g.user:
        return redirect(url_for("login"))
    return render_template("error.html", code=403,
                           msg="شما به این بخش دسترسی ندارید."), 403


@app.errorhandler(404)
def _404(_e):
    if not getattr(g, "user", None):
        g.user = current_user()
    if not g.user:
        return redirect(url_for("login"))
    return render_template("error.html", code=404, msg="صفحه یافت نشد."), 404


@app.errorhandler(413)
def _413(_e):
    if not getattr(g, "user", None):
        g.user = current_user()
    if not g.user:
        return redirect(url_for("login"))
    msg = (f"حجم فایل ارسالی از سقف مجاز ({max_upload_mb()} مگابایت) بیشتر است. "
           "فایل کوچک‌تری انتخاب کنید یا از مدیر سامانه بخواهید در «تنظیمات» "
           "سقف حجم آپلود را افزایش دهد. اگر این پیام از وب‌سرور (nginx) آمده است "
           "باید مقدار client_max_body_size در تنظیمات nginx افزایش یابد.")
    return render_template("error.html", code=413, msg=msg), 413


with app.app_context():
    init_db()

if __name__ == "__main__":
    port = int(os.environ.get("PORT", "5000"))
    print(f"➡ در حال اجرا: http://127.0.0.1:{port}   |   ورود مدیر: admin / admin123")
    app.run(host="0.0.0.0", port=port, debug=False)

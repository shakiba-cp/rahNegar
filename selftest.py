# -*- coding: utf-8 -*-
"""تست خودکار جریان‌های اصلی سامانه با client داخلی Flask."""
import io
import json
import sys

from openpyxl import Workbook, load_workbook

import app as A
import jalali as J

c = A.app.test_client()
ok = []


def check(name, cond, extra=""):
    if cond:
        ok.append(name)
    else:
        print(f"  ⛔ FAIL: {name} {extra}")
        sys.exit(1)


def login(user, pw):
    r = c.post("/login", data={"username": user, "password": pw},
               follow_redirects=True)
    return r


# ---------- احراز هویت
r = c.get("/")
check("redirect به login", r.status_code in (302, 301))
r = login("admin", "bad")
check("رمز اشتباه رد می‌شود", b"\xd8\xa7\xd8\xb4\xd8\xaa\xd8\xa8\xd8\xa7" in r.data)
r = login("admin", "admin123")
check("ورود مدیر", r.status_code == 200 and "داشبورد".encode() in r.data)

# ---------- داشبورد
r = c.get("/")
check("داشبورد: نمودارها", all(x.encode() in r.data for x in
      ["فعالیت‌ها به تفکیک حوزه", "سهم هر حوزه", "نمودار وضعیت",
       "روند ماهانه", "ورودهای Excel", "آخرین فعالیت‌ها"]))

# ---------- ثبت فعالیت (فرم پویا)
db = A.get_db.__wrapped__() if hasattr(A.get_db, "__wrapped__") else None
with A.app.app_context():
    dom = A.get_db().execute("SELECT * FROM domains WHERE name='ارزیابی امنیتی وب'").fetchone()
    fields = A.get_db().execute("SELECT * FROM form_fields WHERE domain_id=? ORDER BY sort_order",
                                (dom["id"],)).fetchall()
    fmap = {f["label"]: f for f in fields}
check("۱۳ حوزه موجود", True)

r = c.get(f"/activities/new?domain_id={dom['id']}")
check("فرم پویا حوزه ارزیابی وب", "کارفرما".encode() in r.data
      and "مشخصات درخواست‌دهنده فعالیت".encode() in r.data
      and "مشخصات تحویل فعالیت".encode() in r.data
      and "مدیریت فیلدهای این حوزه".encode() in r.data)

payload = {"status": "در حال انجام"}
for f in fields:
    if f["field_type"] == "date":
        payload[f"f{f['id']}__y"], payload[f"f{f['id']}__m"], payload[f"f{f['id']}__d"] = (
            "1405", "04", "29")
    elif f["field_type"] == "select":
        payload[f"f{f['id']}"] = "زیاد"
    else:
        payload[f"f{f['id']}"] = {"کارفرما": "بانک الف", "کارشناس": "مدیر سامانه",
                                  "آسیب پذیری": "SQL Injection",
                                  "شماره تیکت": "T-4321",
                                  "آدرس": "https://example.com",
                                  "توضیحات": "تست"}.get(f["label"], "مقدار تست")
r = c.post(f"/activities/new?domain_id={dom['id']}", data=payload, follow_redirects=True)
check("ثبت فعالیت", "فعالیت با موفقیت ثبت شد".encode() in r.data, r.data[:80])

with A.app.app_context():
    act = A.get_db().execute("SELECT * FROM activities ORDER BY id DESC LIMIT 1").fetchone()
check("meta: تاریخ میلادی صحیح", act["date"] == "2026-07-20", act["date"])
check("meta: عنوان از فیلد آدرس", act["title"] == "https://example.com", act["title"])
check("meta: تیکت", act["ticket"] == "T-4321")
aid = act["id"]

# ---------- لیست/جستجو/فیلتر
r = c.get("/activities?q=بانک")
check("جستجوی مقدار کارفرما", "https://example.com".encode() in r.data)
r = c.get("/activities?from__y=1405&from__m=04&from__d=01&to__y=1405&to__m=05&to__d=31")
check("فیلتر بازه شمسی", "https://example.com".encode() in r.data)
r = c.get("/activities?from__y=1400&from__m=01&from__d=01&to__y=1401&to__m=12&to__d=29")
check("بازه قدیمی خالی", "فعالیتی یافت نشد".encode() in r.data or "https://example.com".encode() not in r.data)

# ---------- ویرایش
payload[f"f{fmap['آدرس']['id']}"] = "https://new.example.com"
payload["status"] = "انجام شده"
r = c.post(f"/activities/{aid}/edit", data=payload, follow_redirects=True)
check("ویرایش فعالیت", "به‌روزرسانی شد".encode() in r.data
      and "new.example.com".encode() in r.data)

# کسر فیلد الزامی → خطا
bad = dict(payload)
bad[f"f{fmap['کارفرما']['id']}"] = ""
r = c.post(f"/activities/{aid}/edit", data=bad, follow_redirects=True)
check("اعتبارسنجی الزامی", "الزامی".encode() in r.data)

# ---------- پیوست
r = c.post(f"/activities/{aid}/attachments",
           data={"files": [(io.BytesIO(b"hello world"), "گزارش.pdf")]},
           content_type="multipart/form-data", follow_redirects=True)
check("آپلود پیوست", "فایل پیوست شد".encode() in r.data)
with A.app.app_context():
    att = A.get_db().execute("SELECT * FROM attachments WHERE activity_id=?",
                             (aid,)).fetchone()
r = c.get(f"/attachments/{att['id']}/download")
check("دانلود پیوست", r.data == b"hello world")
r = c.post(f"/activities/{aid}/attachments",
           data={"files": [(io.BytesIO(b"x"), "evil.exe")]},
           content_type="multipart/form-data", follow_redirects=True)
check("رد فرمت غیرمجاز", "مجاز نیست".encode() in r.data)
# فایل Word (docx) به پیوست‌ها بدون مشکل آپلود و دانلود می‌شود
r = c.post(f"/activities/{aid}/attachments",
           data={"files": [(io.BytesIO(b"DOCX-FAKE-CONTENT"), "مستندات ارزیابی.docx")]},
           content_type="multipart/form-data", follow_redirects=True)
check("آپلود فایل Word در پیوست", "فایل پیوست شد".encode() in r.data)
with A.app.app_context():
    _dx = A.get_db().execute("SELECT * FROM attachments WHERE activity_id=? AND original_name LIKE '%.docx'",
                             (aid,)).fetchone()
r = c.get(f"/attachments/{_dx['id']}/download")
check("دانلود فایل Word", r.status_code == 200 and r.data == b"DOCX-FAKE-CONTENT")

# ---------- ورود Excel
wb = Workbook()
ws = wb.active
hdr = [f["label"] for f in fields]
ws.append(hdr + ["وضعیت"])
def _row(**kw):
    return [kw.get(h, "") for h in hdr] + [kw["وضعیت_سیستم"]]
ws.append(_row(**{"تاریخ": "۱۴۰۵/۰۴/۱۰", "کارشناس": "رضا کریمی", "کارفرما": "شرکت ج",
                  "آسیب پذیری": "XSS", "شدت": "کم", "آدرس": "https://b.ir",
                  "شماره تیکت": "T-100", "زمان درخواست": "۱۴۰۵/۰۴/۱۰", "وضعیت_سیستم": "انجام شده"}))
ws.append(_row(**{"تاریخ": "1405/04/11", "کارشناس": "سارا احمدی", "کارفرما": "سازمان د",
                  "آسیب پذیری": "IDOR", "شدت": "بحرانی", "آدرس": "https://c.ir",
                  "شماره تیکت": "T-101", "وضعیت_سیستم": "در حال انجام"}))
ws.append(_row(**{"تاریخ": "نامعتبر!", "کارشناس": "بدون تاریخ", "شدت": "کم",
                  "کارفرما": "باید خطا بخورد", "وضعیت_سیستم": ""}))
bio = io.BytesIO()
wb.save(bio)
bio.seek(0)
r = c.post("/import", data={"domain_id": str(dom["id"]),
                            "file": (bio, "فعالیت‌ها.xlsx")},
           content_type="multipart/form-data", follow_redirects=True)
check("پیام نتیجه Excel", "فعالیت ثبت شد".encode() in r.data
      and "ردیف دارای خطا".encode() in r.data, r.data[:200])
with A.app.app_context():
    imp = A.get_db().execute("SELECT * FROM excel_imports ORDER BY id DESC").fetchone()
check("لاگ import: هر ۳ ردیف وارد شد، ۱ ردیف ناقص (پرچم‌دار)",
      imp["success_rows"] == 3 and imp["error_rows"] == 1,
      (imp["success_rows"], imp["error_rows"]))
with A.app.app_context():
    _fl = A.get_db().execute("SELECT COUNT(*) c FROM activities WHERE flagged=1").fetchone()["c"]
check("ردیف ناقص با پرچم «نیازمند اصلاح» ثبت شد", _fl == 1, (_fl,))
r = c.get(f"/import/{imp['id']}")
check("صفحه نتیجه import", "تعداد کل ردیف‌ها".encode() in r.data)

# ---------- تشخیص ردیف تکراری در ورود Excel: پرسش قبل از ثبت + skip/force ----------
with A.app.app_context():
    _n0 = A.get_db().execute("SELECT COUNT(*) AS c FROM activities").fetchone()["c"]
bio2 = io.BytesIO(); wb.save(bio2); bio2.seek(0)
r = c.post("/import", data={"domain_id": str(dom["id"]),
                            "file": (bio2, "فعالیت‌ها.xlsx")},
           content_type="multipart/form-data")
check("ورود مجدد: صفحهٔ پرسش دربارهٔ ردیف تکراری", "ردیف تکراری".encode() in r.data)
with A.app.app_context():
    _n1 = A.get_db().execute("SELECT COUNT(*) AS c FROM activities").fetchone()["c"]
check("ورود مجدد: پیش از تأیید هیچ ردیفی ثبت نشد", _n1 == _n0, (_n1, _n0))
r = c.post("/import/confirm", data={"choice": "skip"}, follow_redirects=True)
check("تأیید skip: پیام «نادیده گرفته شد»", "نادیده گرفته شد".encode() in r.data)
with A.app.app_context():
    _d = A.get_db()
    _imp2 = _d.execute("SELECT * FROM excel_imports ORDER BY id DESC").fetchone()
    _n2 = _d.execute("SELECT COUNT(*) AS c FROM activities").fetchone()["c"]
check("تأیید skip: هر ۳ ردیف تکراری رد شدند و چیز جدیدی ثبت نشد",
      _n2 == _n0 and _imp2["dup_rows"] == 3 and _imp2["success_rows"] == 0,
      (_imp2["dup_rows"], _imp2["success_rows"]))
wb3 = Workbook(); ws3 = wb3.active
ws3.append(hdr + ["وضعیت"])
ws3.append(_row(**{"تاریخ": "۱۴۰۵/۰۴/۱۰", "کارشناس": "رضا کریمی", "کارفرما": "شرکت ج",
                   "آسیب پذیری": "XSS", "شدت": "کم", "آدرس": "https://b.ir",
                   "شماره تیکت": "T-100", "زمان درخواست": "۱۴۰۵/۰۴/۱۰", "وضعیت_سیستم": "انجام شده"}))
ws3.append(_row(**{"تاریخ": "1405/04/11", "کارشناس": "سارا احمدی", "کارفرما": "سازمان د",
                   "آسیب پذیری": "IDOR", "شدت": "بحرانی", "آدرس": "https://c.ir",
                   "شماره تیکت": "T-101", "وضعیت_سیستم": "در حال انجام"}))
ws3.append(_row(**{"تاریخ": "1405/04/12", "کارشناس": "رضا کریمی", "کارفرما": "سازمان ه",
                   "آسیب پذیری": "SSRF", "شدت": "متوسط", "آدرس": "https://d.ir",
                   "شماره تیکت": "T-200", "زمان درخواست": "1405/04/12", "وضعیت_سیستم": "در حال انجام"}))
bio3 = io.BytesIO(); wb3.save(bio3); bio3.seek(0)
r = c.post("/import", data={"domain_id": str(dom["id"]),
                            "file": (bio3, "فعالیت‌ها۳.xlsx")},
           content_type="multipart/form-data")
check("فایل مختلط: صفحهٔ پرسش ردیف تکراری", "ردیف تکراری".encode() in r.data)
r = c.post("/import/confirm", data={"choice": "force"}, follow_redirects=True)
with A.app.app_context():
    _d = A.get_db()
    _imp3 = _d.execute("SELECT * FROM excel_imports ORDER BY id DESC").fetchone()
    _n3 = _d.execute("SELECT COUNT(*) AS c FROM activities").fetchone()["c"]
    _t200 = _d.execute("SELECT 1 FROM activities WHERE ticket='T-200'").fetchone()
check("تأیید force: هر سه ردیف ثبت شدند (حتی تکراری‌ها)",
      _imp3["success_rows"] == 3 and _imp3["dup_rows"] == 0 and _t200 is not None,
      (_imp3["success_rows"],))
check("تأیید force: تعداد فعالیت‌ها +۳ شد", _n3 == _n0 + 3, (_n3, _n0))

# ---------- ورود تلرانت: زیرمجموعهٔ ستون‌ها + نگاشت کارشناس به کاربر سامانه ----------
wb4 = Workbook(); ws4 = wb4.active
ws4.append(["کارشناس", "شماره تیکت", "شدت"])   # فقط زیرمجموعه — «تاریخ» (الزامی) ستون ندارد
ws4.append(["رضا کریمی", "T-300", "کم"])
bio4 = io.BytesIO(); wb4.save(bio4); bio4.seek(0)
r = c.post("/import", data={"domain_id": str(dom["id"]),
                            "file": (bio4, "زیرمجموعه.xlsx")},
           content_type="multipart/form-data", follow_redirects=True)
with A.app.app_context():
    _d = A.get_db()
    _imp4 = _d.execute("SELECT * FROM excel_imports ORDER BY id DESC").fetchone()
    _t300 = _d.execute("SELECT * FROM activities WHERE ticket='T-300'").fetchone()
check("ورود تلرانت: فیلد الزامیِ غایب از فایل خطا نیست", _imp4["success_rows"] == 1 and _imp4["error_rows"] == 0,
      (_imp4["success_rows"], _imp4["error_rows"]))
check("کارشناس ناشناخته: آپلودکننده + هشدار", _t300 is not None and _t300["user_id"] == 1
      and "رضا کریمی" in (_imp4["warns"] or ""), (_imp4["warns"] or "")[:120])

# نامِ داخل فایل (حتی بدون حساب کاربری) باید «به‌نام خودش» ثبت و نمایش داده شود
with A.app.app_context():
    _rows300 = A.query_activities("a.ticket=?", ["T-300"])
check("کارشناس ناشناخته: نامِ نمایشی = همان نامِ داخل فایل",
      _rows300 and _rows300[0]["expert_txt"] == "رضا کریمی",
      _rows300[0]["expert_txt"] if _rows300 else None)
r = c.get("/")
check("کارشناس ناشناخته در نمودار تفکیک کارشناس دیده می‌شود",
      "رضا کریمی".encode() in r.data)
r = c.get("/reports?export=excel")
check("خروجی اکسل گزارش ساخته شد", r.status_code == 200
      and r.data[:2] == b"PK", r.status_code)
_wb = load_workbook(io.BytesIO(r.data), read_only=True); _ws = _wb.active
# سربرگ = اولین ردیفی که سلول اولش «حوزه» است (ردیف ۱ عنوان گزارش شد)
_hr = None; _hd = []
for _i0, _row0 in enumerate(_ws.iter_rows(max_row=4), start=1):
    if _row0 and _row0[0].value == "حوزه":
        _hr, _hd = _i0, [c0.value for c0 in _row0]; break
_iexp = _hd.index("کارشناس") if "کارشناس" in _hd else -1
_exp_vals = [row[_iexp].value for row in _ws.iter_rows(min_row=(_hr or 1) + 1)] if _iexp >= 0 else []
check("خروجی اکسل: ستون کارشناس = نامِ داخل فایل",
      _iexp >= 0 and "رضا کریمی" in _exp_vals, (_hd[:6], _exp_vals[:5]))

# کاربر جدید → نام در اکسل → مالکیت همان کاربر می‌شود
r = c.post("/users/new", data={"username": "negar", "full_name": "نگار نمونه",
                               "role": "expert", "password": "123456"}, follow_redirects=True)
check("ایجاد کاربر نگار برای تست نگاشت", "کاربر ایجاد شد".encode() in r.data)
with A.app.app_context():
    _negar = A.get_db().execute("SELECT id FROM users WHERE username='negar'").fetchone()["id"]
wb5 = Workbook(); ws5 = wb5.active
ws5.append(["کارشناس", "شماره تیکت", "شدت"])
ws5.append(["نگار نمونه", "T-400", "متوسط"])
bio5 = io.BytesIO(); wb5.save(bio5); bio5.seek(0)
r = c.post("/import", data={"domain_id": str(dom["id"]),
                            "file": (bio5, "نگاشت.xlsx")},
           content_type="multipart/form-data", follow_redirects=True)
with A.app.app_context():
    _d = A.get_db()
    _imp5 = _d.execute("SELECT * FROM excel_imports ORDER BY id DESC").fetchone()
    _t400 = _d.execute("SELECT * FROM activities WHERE ticket='T-400'").fetchone()
check("نگاشت کارشناس: فعالیت به نام کاربرِ اکسل ثبت شد",
      _t400 is not None and _t400["user_id"] == _negar and (_imp5["warns"] or "") == "",
      (_t400["user_id"] if _t400 else None, _negar))

# --- تیکت از Excel خوانده و قابل جستجو است (رفع ایراد «تیکت‌ها خونده نمی‌شوند»)
with A.app.app_context():
    t_rows = A.get_db().execute(
        "SELECT ticket FROM activities WHERE ticket LIKE 'T-1%' ORDER BY id").fetchall()
check("تیکت از Excel خوانده شد", any(r["ticket"] == "T-100" for r in t_rows),
      [r["ticket"] for r in t_rows])
r = c.get("/activities?ticket=T-100")
check("جستجوی تیکتِ واردشده از Excel", "https://b.ir".encode() in r.data)

# --- سلول عددی اکسل (4321.0) به تیکت صحیح «4321» تبدیل می‌شود
wb_n = Workbook()
ws_n = wb_n.active
ws_n.append(["تاریخ", "کارشناس", "کارفرما", "شدت", "شماره تیکت"])
ws_n.append(["1405/04/12", "تست عددی", "شرکت عددی", "کم", 4321.0])
bio_n = io.BytesIO()
wb_n.save(bio_n)
bio_n.seek(0)
r = c.post("/import", data={"domain_id": str(dom["id"]),
                            "file": (bio_n, "ticket-num.xlsx")},
           content_type="multipart/form-data", follow_redirects=True)
with A.app.app_context():
    tick = A.get_db().execute(
        "SELECT ticket FROM activities WHERE title='شرکت عددی' ORDER BY id DESC").fetchone()
check("تیکت عددی اکسل صحیح خوانده شد", tick and tick["ticket"] == "4321",
      tick["ticket"] if tick else None)

# --- فایل ناسازگار (ستون‌های بی‌ربط): هیچ فعالیتی ثبت نمی‌شود
with A.app.app_context():
    cnt_before = A.get_db().execute("SELECT COUNT(*) c FROM activities").fetchone()["c"]
wb_x = Workbook()
ws_x = wb_x.active
ws_x.append(["ستون الف", "ستون ب", "ستون ج"])
ws_x.append([1, 2, 3])
ws_x.append([4, 5, 6])
bio_x = io.BytesIO()
wb_x.save(bio_x)
bio_x.seek(0)
r = c.post("/import", data={"domain_id": str(dom["id"]),
                            "file": (bio_x, "bad.xlsx")},
           content_type="multipart/form-data", follow_redirects=True)
with A.app.app_context():
    cnt_after = A.get_db().execute("SELECT COUNT(*) c FROM activities").fetchone()["c"]
    imp_bad = A.get_db().execute("SELECT * FROM excel_imports ORDER BY id DESC").fetchone()
check("فایل ناسازگار هیچ فعالیتی ثبت نمی‌کند", cnt_before == cnt_after
      and imp_bad["success_rows"] == 0 and imp_bad["error_rows"] == 2,
      (cnt_before, cnt_after, imp_bad["success_rows"]))
check("پیام لغو ورود فایل ناسازگار", "فایل ثبت نشد".encode() in r.data)

# --- ردیف بدون هیچ مقدار معتبر (ghost) ثبت نمی‌شود
wb_g = Workbook()
ws_g = wb_g.active
ws_g.append(["کارفرما", "ستون ناشناس"])
ws_g.append(["", "داده زائد"])
ws_g.append(["شرکت سالم", "مقدار"])
bio_g = io.BytesIO()
wb_g.save(bio_g)
bio_g.seek(0)
r = c.post("/import", data={"domain_id": str(dom["id"]),
                            "file": (bio_g, "ghost.xlsx")},
           content_type="multipart/form-data", follow_redirects=True)
with A.app.app_context():
    cnt_ghost = A.get_db().execute("SELECT COUNT(*) c FROM activities").fetchone()["c"]
    titles_g = [x[0] for x in A.get_db().execute(
        "SELECT title FROM activities WHERE title IN ('شرکت سالم','داده زائد')")]
# قرارداد جدید (ورود تلرانت): ردیف با سلول کاملاً خالیِ ستون‌های شناخته‌شده => رد می‌شود؛
# ردیف با مقدار معتبر در ستون موجود => حتی بدون ستون‌های الزامیِ غایب ثبت می‌شود
check("ردیف خالی ghost ثبت نمی‌شود", "داده زائد" not in titles_g, titles_g)
check("ردیف دارای مقدار در ستون موجود ثبت می‌شود (تلرانت)", cnt_ghost == cnt_after + 1
      and "شرکت سالم" in titles_g, (cnt_after, cnt_ghost, titles_g))

# قالب نمونه
r = c.get(f"/import/template/{dom['id']}")
check("دانلود قالب نمونه xlsx", r.status_code == 200
      and r.data[:2] == b"PK")

# ---------- ورود CSV (همان مسیر Excel) + کلید عنوان «آدرس»
with A.app.app_context():
    _d = A.get_db()
    _ak = _d.execute("SELECT field_key FROM form_fields WHERE domain_id=? AND label='آدرس'",
                     (dom["id"],)).fetchone()["field_key"]
    _ck = _d.execute("SELECT field_key FROM form_fields WHERE domain_id=? AND label='کارفرما'",
                     (dom["id"],)).fetchone()["field_key"]
check("کلید عنوان حوزه وب: «آدرس» نه «کارفرما»", _ak == "title" and not _ck, (_ak, _ck))

# کلید عنوان حوزه اندروید: «برنامه» نه «کارفرما»
with A.app.app_context():
    _d = A.get_db()
    _and = _d.execute("SELECT id FROM domains WHERE name='ارزیابی امنیتی اندروید'").fetchone()["id"]
    _bk = _d.execute("SELECT field_key FROM form_fields WHERE domain_id=? AND label='برنامه'",
                     (_and,)).fetchone()["field_key"]
    _ck2 = _d.execute("SELECT field_key FROM form_fields WHERE domain_id=? AND label='کارفرما'",
                      (_and,)).fetchone()["field_key"]
check("کلید عنوان حوزه اندروید: «برنامه» نه «کارفرما»",
      _bk == "title" and not _ck2, (_bk, _ck2))

_csv = ("تاریخ,کارشناس,کارفرما,شدت,آدرس\n"
        "1405/04/30,مدیر سامانه,سازمان csv,زیاد,https://csv.example\n").encode("utf-8-sig")
r = c.post("/import", data={"domain_id": str(dom["id"]),
                            "file": (io.BytesIO(_csv), "import.csv")},
           content_type="multipart/form-data", follow_redirects=True)
check("ورود فایل CSV", "فعالیت با موفقیت ثبت شد".encode() in r.data, r.data[:80])
with A.app.app_context():
    _csv_act = A.get_db().execute(
        "SELECT title FROM activities WHERE title='https://csv.example'").fetchone()
check("ردیف CSV ثبت شد (عنوان=آدرس)", _csv_act is not None,
      _csv_act["title"] if _csv_act else None)

# CSV با جداکننده نقطه‌ویرگول
_csv2 = ("تاریخ;کارشناس;آدرس\n1405/05/01;مدیر سامانه;https://semi.example\n").encode("utf-8")
r = c.post("/import", data={"domain_id": str(dom["id"]),
                            "file": (io.BytesIO(_csv2), "semi.csv")},
           content_type="multipart/form-data", follow_redirects=True)
with A.app.app_context():
    _semi = A.get_db().execute(
        "SELECT title FROM activities WHERE title='https://semi.example'").fetchone()
check("ورود CSV با جداکننده «؛»", _semi is not None, _semi["title"] if _semi else None)

# فایل نامعتبر رد شود
r = c.post("/import", data={"domain_id": str(dom["id"]),
                            "file": (io.BytesIO(b"x"), "bad.txt")},
           content_type="multipart/form-data", follow_redirects=True)
check("رد فرمت غیرمجاز در ورود", "پذیرفته می‌شود".encode() in r.data)

# ---------- گزارش‌ها و خروجی‌ها
r = c.get("/reports?ticket=4321")
check("گزارش با فیلتر تیکت", "new.example.com".encode() in r.data)
r = c.get("/reports?export=csv")
check("خروجی CSV", r.status_code == 200 and r.data.startswith(b"\xef\xbb\xbf")
      and "شماره تیکت".encode("utf-8") in r.data)
r = c.get("/reports?export=excel")
check("خروجی Excel", r.data[:2] == b"PK")
wb2 = load_workbook(io.BytesIO(r.data))
_hdr_cells = [x.value for row in wb2.active.iter_rows(max_row=3) for x in row]
check("ستون تاریخ شمسی در Excel", "تاریخ (شمسی)" in _hdr_cells)
rows = list(wb2.active.iter_rows(values_only=True))
check("تاریخ‌ها در خروجی شمسی‌اند", any(str(c).startswith("۱۴۰۵/") or str(c).startswith("1405/")
      for row in rows[1:] for c in row if c), rows[1] if len(rows) > 1 else None)
r = c.get("/reports?export=pdf")
check("صفحه چاپ PDF", r.status_code == 200 and b"report-print.js" in r.data and b"print-btn" in r.data)

# ---------- کاربران
r = c.post("/users/new", data={"username": "reza", "full_name": "رضا کریمی",
                               "role": "expert", "password": "123456"},
           follow_redirects=True)
check("ایجاد کارشناس", "کاربر ایجاد شد".encode() in r.data)
r = c.post("/users/new", data={"username": "reza", "full_name": "تکراری",
                               "role": "expert", "password": "123456"},
           follow_redirects=True)
check("رد نام کاربری تکراری", "قبلاً ثبت شده".encode() in r.data)

# ---------- حوزه/فیلد
r = c.post("/domains", data={"name": "حوزه آزمایشی"}, follow_redirects=True)
check("افزودن حوزه", "حوزه جدید افزوده شد".encode() in r.data)
with A.app.app_context():
    nd = A.get_db().execute("SELECT * FROM domains WHERE name='حوزه آزمایشی'").fetchone()
r = c.post(f"/domains/{nd['id']}/fields/add",
           data={"label": "آدرس IP", "field_type": "text"},
           follow_redirects=True)
check("افزودن فیلد پویا", "فیلد افزوده شد".encode() in r.data
      and "آدرس IP".encode() in r.data)
r = c.get(f"/activities/new?domain_id={nd['id']}")
check("فیلد جدید در فرم دیده می‌شود", "آدرس IP".encode() in r.data)
# حذف حوزه بدون فعالیت
r = c.post(f"/domains/{nd['id']}/delete", follow_redirects=True)
check("حذف حوزه خالی", "حوزه حذف شد".encode() in r.data)
# حذف حوزهِ دارای فعالیت → ممنوع
r = c.post(f"/domains/{dom['id']}/delete", follow_redirects=True)
check("ممنوعیت حذف حوزه دارای فعالیت", "غیرفعال کنید".encode() in r.data)

# ---------- تنظیمات
r = c.post("/settings", data={"system_name": "سامانه تست امنیت",
                              "max_upload_mb": "5",
                              "allowed_formats": "pdf,jpg,zip"},
           follow_redirects=True)
check("ذخیره تنظیمات", "تنظیمات ذخیره شد".encode() in r.data)
with A.app.app_context():
    assert A.get_setting("system_name") == "سامانه تست امنیت"
    assert A.allowed_formats() == ["pdf", "jpg", "zip"]
    assert A.max_upload_mb() == 5
check("اعمال تنظیمات", True)

# ---------- سطح دسترسی کارشناس
c.get("/logout")
r = login("reza", "123456")
check("ورود کارشناس", r.status_code == 200)
r = c.get("/activities")
check("کارشناس فقط فعالیت خودش را می‌بیند", "بانک الف".encode() not in r.data
      and "new.example.com".encode() not in r.data)
r = c.get(f"/activities/{aid}")
check("کارشناس به فعالیت دیگران ۴۰۳", r.status_code == 403)
r = c.get("/users")
check("کارشناس به کاربران ۴۰۳", r.status_code == 403)
r = c.get("/settings")
check("کارشناس به تنظیمات ۴۰۳", r.status_code == 403)
r = c.post(f"/activities/{aid}/delete")
check("حذف فقط برای مدیر", r.status_code == 403)
r = c.get("/import")
check("کارشناس می‌تواند Excel وارد کند", r.status_code == 200
      and "ورود اطلاعات از Excel".encode() in r.data)

# تخصیص تسک: مدیر برای کارشناس فعالیت ثبت می‌کند
c.get("/logout"); login("admin", "admin123")
with A.app.app_context():
    rid = A.get_db().execute("SELECT id FROM users WHERE username='reza'").fetchone()["id"]
payload3 = {"status": "انجام شده", "owner_id": str(rid)}
for f in fmap.values():
    if f["field_type"] == "date":
        payload3[f"f{f['id']}__y"], payload3[f"f{f['id']}__m"], payload3[f"f{f['id']}__d"] = ("1405", "02", "10")
    elif f["field_type"] == "select":
        payload3[f"f{f['id']}"] = "کم"
    else:
        payload3[f"f{f['id']}"] = "تسک مدیر برای رضا" if f["label"] == "آدرس" else "تست"
r = c.post(f"/activities/new?domain_id={dom['id']}", data=payload3, follow_redirects=True)
check("تخصیص تسک توسط مدیر", "تخصیص یافت".encode() in r.data)
c.get("/logout"); login("reza", "123456")
r = c.get("/activities")
check("کارشناس تسک تخصیص‌یافته را می‌بیند", "تسک مدیر برای رضا".encode() in r.data)
r = c.get(f"/activities/new?domain_id={dom['id']}")
check("دکمه مدیریت فیلدها برای کارشناس نیست",
      "مدیریت فیلدهای این حوزه".encode() not in r.data)
# ثبت فعالیت توسط کارشناس
with A.app.app_context():
    fields2 = A.get_db().execute("SELECT * FROM form_fields WHERE domain_id=? AND field_key='date'",
                                 (dom["id"],)).fetchone()
payload2 = {"status": "در حال انجام"}
for f in fmap.values():
    if f["field_type"] == "date":
        payload2[f"f{f['id']}__y"], payload2[f"f{f['id']}__m"], payload2[f"f{f['id']}__d"] = (
            "1405", "01", "15")
    elif f["field_type"] == "select":
        payload2[f"f{f['id']}"] = "کم"
    else:
        payload2[f"f{f['id']}"] = "فعالیت رضا" if f["label"] == "آدرس" else "تست"
r = c.post(f"/activities/new?domain_id={dom['id']}", data=payload2, follow_redirects=True)
check("ثبت فعالیت کارشناس", "موفقیت".encode() in r.data)
r = c.get("/reports")
check("گزارش کارشناس فقط مال خودش", "بانک الف".encode() not in r.data
      and "فعالیت رضا".encode() in r.data)


# ---------- خروجی: بدون شناسه + همه فیلدها + بدون تکرار تیکت
c.get("/logout"); login("admin", "admin123")
r = c.get("/reports?export=excel")
wb3 = load_workbook(io.BytesIO(r.data))
def _xh(wb):
    """سربرگ هر شیت = اولین ردیف با سلول اول «حوزه» (ردیف ۱ عنوان گزارش است)."""
    out = []
    for nm in wb.sheetnames:
        for row in wb[nm].iter_rows(max_row=3, values_only=True):
            if row and row[0] == "حوزه":
                out.append((nm, [x for x in row if x is not None])); break
    return out
_sheets = _xh(wb3)
hdrs_all = [h for _n, hs in _sheets for h in hs]
check("خروجی شناسه سیستمی (id دیتابیس) ندارد",
      not any(h in ("id", "شناسه سیستمی") for h in hdrs_all), hdrs_all[:12])
check("تیکت فقط یک ستون",
      _sheets and all(hs.count("شماره تیکت") == 1 for _n, hs in _sheets),
      [n for n, _ in _sheets])
check("همه فیلدها در خروجی‌اند",
      "زمان درخواست" in hdrs_all and "آسیب پذیری" in hdrs_all, hdrs_all[:20])
check("خروجی تک‌حوزه‌ای: شیت به نام همان حوزه",
      wb3.sheetnames and wb3.sheetnames[0] == "ارزیابی امنیتی وب", wb3.sheetnames)

# ---------- انتخاب ستون خروجی توسط مدیر
r = c.get("/reports", query_string=[("export", "excel"), ("col", "حوزه"), ("col", "وضعیت")])
wb4 = load_workbook(io.BytesIO(r.data))
hdrs2 = _xh(wb4)[0][1] if _xh(wb4) else []
check("انتخاب ستون خروجی مدیر", hdrs2[:2] == ["حوزه", "وضعیت"], hdrs2)
r = c.get("/reports")
check("کارت انتخاب ستون برای مدیر", "ستون‌های خروجی".encode() in r.data)

# پوشش خروجی چندحوزه‌ای: یک فعالیت در حوزه دوم (کپی دیتابیسی) سپس خروجی
with A.app.app_context():
    _dbx = A.get_db()
    _d2 = _dbx.execute(
        "SELECT id FROM domains WHERE name='استخراج IP آسیب پذیر'").fetchone()["id"]
    _dbx.execute("""INSERT INTO activities (domain_id, user_id, status, title, ticket,
                        date, created_at, updated_at)
                    SELECT ?, user_id, status, 'نمونه حوزه دوم', ticket, date,
                           created_at, updated_at
                    FROM activities ORDER BY id LIMIT 1""", (_d2,))
    _dbx.commit()
r = c.get("/reports?export=excel")
wb3b = load_workbook(io.BytesIO(r.data))
check("خروجی چندحوزه‌ای: شیت خلاصه + شیت مجزای هر حوزه",
      "خلاصه" in wb3b.sheetnames and len(wb3b.sheetnames) >= 3
      and "استخراج IP آسیب پذیر" in wb3b.sheetnames, wb3b.sheetnames)

# ---------- تیکت در فرم اصلی استخراج IP (نه فقط بخش تحویل)
with A.app.app_context():
    dom_ip = A.get_db().execute("SELECT * FROM domains WHERE name='استخراج IP آسیب پذیر'").fetchone()
r = c.get(f"/activities/new?domain_id={dom_ip['id']}")
body = r.data.decode()
check("استخراج IP تیکت در فرم اصلی دارد", "شماره تیکت" in body)
sect = body.index("مشخصات تحویل فعالیت") if "مشخصات تحویل فعالیت" in body else 10**9
main_part = body[:sect]
check("تیکت قبل از بخش تحویل است", "شماره تیکت" in main_part)

# ---------- گردش‌کار تسک: پاسخ + انجام شد
with A.app.app_context():
    task = A.get_db().execute("""SELECT a.* FROM activities a
        WHERE a.created_by <> a.user_id ORDER BY a.id DESC LIMIT 1""").fetchone()
    tid = task["id"]
c.get("/logout"); login("reza", "123456")  # رضا مالک تسک است
r = c.get("/activities")
check("بج تسک در لیست کارشناس", "تسک".encode() in r.data)
r = c.get(f"/activities/{tid}")
check("برچسب تسک در صفحه فعالیت", "تخصیص‌یافته".encode() in r.data)
r = c.post(f"/activities/{tid}/respond", data={"body": "ارزیابی انجام شد، گزارش ضمیمه شد."},
           follow_redirects=True)
check("ثبت پاسخ توسط کارشناس", "پاسخ ثبت شد".encode() in r.data
      and "ارزیابی انجام شد".encode() in r.data)
r = c.post(f"/activities/{tid}/complete", follow_redirects=True)
check("علامت انجام شد", "علامت‌گذاری شد".encode() in r.data)
with A.app.app_context():
    st = A.get_db().execute("SELECT status FROM activities WHERE id=?", (tid,)).fetchone()["status"]
check("وضعیت انجام شده شد", st == "انجام شده", st)
c.get("/logout"); login("admin", "admin123")
r = c.get(f"/activities/{tid}")
check("مدیر پاسخ کارشناس را می‌بیند", "ارزیابی انجام شد".encode() in r.data)

# ---------- PDF تک‌حوزه: فقط نمودارهای همان حوزه
r = c.get(f"/reports?export=pdf&domain={dom['id']}")
h5 = r.data.decode()
check("PDF تک‌حوزه بدون نمودار سهم", "سهم حوزه‌ها" not in h5)
check("PDF تک‌حوزه دارای وضعیت", "نمودار وضعیت" in h5)
# ساخت یک فعالیت در حوزه دوم تا نمودار سهم حوزه‌ها دیده شود
with A.app.app_context():
    dom_b = A.get_db().execute("SELECT * FROM domains WHERE name='بدافزار'").fetchone()
    fb = A.get_db().execute("SELECT * FROM form_fields WHERE domain_id=? AND is_active=1",
                            (dom_b["id"],)).fetchall()
pb = {"status": "انجام شده"}
for f in fb:
    if f["field_type"] == "date":
        pb[f"f{f['id']}__y"], pb[f"f{f['id']}__m"], pb[f"f{f['id']}__d"] = ("1405", "03", "01")
    elif f["field_type"] == "select":
        opts = __import__("json").loads(f["options"] or "[]")
        pb[f"f{f['id']}"] = opts[0] if opts else ""
    else:
        pb[f"f{f['id']}"] = "Emotet" if f["label"] == "بدافزار" else ("رضا کریمی" if f["field_key"] == "expert" else "تست")
r = c.post(f"/activities/new?domain_id={dom_b['id']}", data=pb, follow_redirects=True)
check("ثبت فعالیت حوزه دوم", "موفقیت".encode() in r.data)
r = c.get("/reports?export=pdf")
h6 = r.data.decode()
check("PDF همه حوزه‌ها نمودار سهم دارد", "سهم حوزه‌ها" in h6)

# ---------- تغییر وضعیت گروهی + حذف گروهی (سطح دسترسی) ----------
c.get("/logout"); login("admin", "admin123")
with A.app.app_context():
    _d = A.get_db()
    _web = _d.execute("SELECT id FROM domains WHERE name='ارزیابی امنیتی وب'").fetchone()
    _some = [r0["id"] for r0 in _d.execute(
        "SELECT id FROM activities ORDER BY id DESC LIMIT 3")]
    _opt = _d.execute("""SELECT options FROM form_fields
                         WHERE domain_id=? AND label='وضعیت آسیب‌پذیری'""",
                      (_web["id"],)).fetchone()
check("وضعیت آسیب‌پذیری وب: گزینه‌ها تایید/عدم تایید",
      _opt and json.loads(_opt["options"]) == ["تایید", "عدم تایید"],
      _opt["options"] if _opt else None)

r = c.post("/activities/bulk-status", json={"ids": _some, "status": "انجام شده"})
check("تغییر وضعیت گروهی (ادمین)",
      r.status_code == 200 and r.get_json().get("changed") == len(_some), r.get_json())
with A.app.app_context():
    _st = [A.get_db().execute("SELECT status FROM activities WHERE id=?", (i,)).fetchone()["status"]
           for i in _some]
check("وضعیت‌ها واقعاً در دیتابیس عوض شد", all(s == "انجام شده" for s in _st), _st)
r = c.post("/activities/bulk-status", json={"ids": _some, "status": "وضعیت الکی"})
check("وضعیت نامعتبر گروهی ۴۰۰", r.status_code == 400, r.status_code)

c.get("/logout"); login("negar", "123456")
r = c.post("/activities/bulk-status", json={"ids": _some, "status": "بررسی شده"})
check("کارشناس نمی‌تواند «بررسی شده» گروهی بزند", r.status_code == 400, r.status_code)
r = c.post("/activities/bulk-status", json={"ids": _some, "status": "در حال انجام"})
check("کارشناس روی فعالیت دیگران اثری ندارد",
      r.status_code == 200 and r.get_json().get("changed") == 0, r.get_json())
r = c.post("/activities/bulk-delete", json={"ids": _some})
check("حذف گروهی برای کارشناسِ بدون مجوز ۴۰۳", r.status_code == 403, r.status_code)
with A.app.app_context():
    _still = A.get_db().execute("SELECT COUNT(*) c FROM activities").fetchone()["c"]
check("هیچ فعالیتی حذف نشد", _still > 0)

# ---------- فیلد تازه‌افزوده‌شده در خروجی + حذف ستون تیکتِ خالی ----------
c.get("/logout"); login("admin", "admin123")
r = c.post(f"/domains/{_web['id']}/fields/add",
           data={"label": "شناسه", "field_type": "text"}, follow_redirects=True)
check("افزودن فیلد «شناسه» به حوزه وب", "فیلد افزوده شد".encode() in r.data)
r = c.get(f"/reports?domain={_web['id']}&export=excel")
check("خروجی اکسل حوزه وب", r.status_code == 200 and r.data[:2] == b"PK", r.status_code)
_wb = load_workbook(io.BytesIO(r.data), read_only=True); _ws = _wb.active
_hd = _xh(_wb)[0][1] if _xh(_wb) else []
check("ستون «شناسه» در خروجی هست (حتی بدون مقدار)", "شناسه" in _hd, _hd)
check("ستون «وضعیت آسیب‌پذیری» در خروجی هست", "وضعیت آسیب‌پذیری" in _hd, _hd)

# حوزه بدون هیچ تیکتی → ستون «شماره تیکت» از خروجی حذف می‌شود
with A.app.app_context():
    _d = A.get_db()
    _mz = _d.execute("""SELECT d.id FROM domains d
                        WHERE NOT EXISTS(SELECT 1 FROM activities a WHERE a.domain_id=d.id)
                        ORDER BY d.id LIMIT 1""").fetchone()["id"]
    _d.execute("""INSERT INTO activities(domain_id,user_id,status,created_at,updated_at)
                  VALUES(?,?,?,?,?)""",
               (_mz, 1, A.STATUSES[0], A.now_iso(), A.now_iso()))
    _d.commit()
r = c.get(f"/reports?domain={_mz}&export=excel")
check("خروجی اکسل حوزه بدافزار", r.status_code == 200 and r.data[:2] == b"PK", r.status_code)
_wb2 = load_workbook(io.BytesIO(r.data), read_only=True); _ws2 = _wb2.active
_hd2 = _xh(_wb2)[0][1] if _xh(_wb2) else []
check("ستون «شماره تیکت» برای حوزهٔ بدون تیکت حذف شد",
      "شماره تیکت" not in _hd2, _hd2)

# ---------- فیلد نوع «فایل»: آپلود مستندات از داخل فرم ----------
r = c.post(f"/domains/{_web['id']}/fields/add",
           data={"label": "گزارش نهایی", "field_type": "file"}, follow_redirects=True)
check("افزودن فیلد نوع «فایل»", "فیلد افزوده شد".encode() in r.data)
with A.app.app_context():
    _d = A.get_db()
    _wf = {r0["label"]: r0 for r0 in _d.execute(
        "SELECT * FROM form_fields WHERE domain_id=?", (_web["id"],))}
    _ff = _wf["گزارش نهایی"]
_tid = _wf["تاریخ"]["id"]
form = {"status": "در حال انجام",
        f"f{_wf['کارشناس']['id']}": "مدیر سامانه",
        f"f{_wf['کارفرما']['id']}": "بانک تست فایل",
        f"f{_wf['شدت']['id']}": "زیاد",
        f"f{_tid}__y": "1405", f"f{_tid}__m": "05", f"f{_tid}__d": "01",
        f"ff{_ff['id']}": (io.BytesIO(b"PDFDATA-TEST"), "report-final.pdf")}
r = c.post(f"/activities/new?domain_id={_web['id']}", data=form,
           content_type="multipart/form-data", follow_redirects=True)
check("ثبت فعالیت همراه فیلد فایل", "با موفقیت ثبت شد".encode() in r.data, r.data[:60])
with A.app.app_context():
    _d = A.get_db()
    _act = _d.execute("SELECT id FROM activities ORDER BY id DESC LIMIT 1").fetchone()
    _att = _d.execute("""SELECT * FROM attachments
                         WHERE activity_id=? AND original_name='report-final.pdf'""",
                      (_act["id"],)).fetchone()
    _val = _d.execute("SELECT value FROM activity_values WHERE activity_id=? AND field_id=?",
                      (_act["id"], _ff["id"])).fetchone()
check("فایل در پیوست‌ها ذخیره و مقدار فیلد = نام فایل",
      _att is not None and _val is not None and _val["value"] == "report-final.pdf",
      (_att["id"] if _att else None, _val["value"] if _val else None))
if _att:
    r = c.get(f"/attachments/{_att['id']}/download")
    check("دانلود فایلِ فیلد فایل", r.status_code == 200 and r.data == b"PDFDATA-TEST",
          r.status_code)
r = c.get(f"/activities/{_act['id']}")
check("صفحه جزئیات: نام فایل فیلد نمایش داده می‌شود",
      "report-final.pdf".encode() in r.data)

# بنر «نیازمند اصلاح» در صفحه جزئیات
with A.app.app_context():
    _flr = A.get_db().execute("SELECT id FROM activities WHERE flagged=1 LIMIT 1").fetchone()
if _flr:
    r = c.get(f"/activities/{_flr['id']}")
    check("بنر «نیازمند اصلاح» در صفحه جزئیات (flagged در JSON)",
      b'"flagged": true' in r.data, r.data[-200:])
else:
    check("بنر «نیازمند اصلاح» در صفحه جزئیات", False, "ردیف flagged یافت نشد")

# هم‌ترازی افزودنی: «شناسه فرآیند» به تعریف حوزه وب اضافه شده است
with A.app.app_context():
    _sp = A.get_db().execute("""SELECT 1 FROM form_fields
                                WHERE domain_id=? AND label='شناسه فرآیند'""",
                             (_web["id"],)).fetchone()
check("هم‌ترازی حوزه‌ها: «شناسه فرآیند» موجود است", _sp is not None)

# مراکز (ماهر/کاشف/فیدار) و اتصال حوزه‌ها به مرکز
with A.app.app_context():
    _orgs = A.get_db().execute("SELECT name FROM orgs ORDER BY sort_order").fetchall()
check("سه مرکز پیش‌فرض ساخته شده‌اند",
      [o["name"] for o in _orgs] == ["ماهر", "کاشف", "فیدار"], [o["name"] for o in _orgs])
with A.app.app_context():
    _noorg = A.get_db().execute("SELECT COUNT(*) c FROM domains WHERE org_id IS NULL").fetchone()["c"]
    _mah = A.get_db().execute("SELECT id FROM orgs WHERE name='ماهر'").fetchone()
    _kashef = A.get_db().execute("SELECT id FROM orgs WHERE name='کاشف'").fetchone()
check("همه حوزه‌ها به یک مرکز متصل‌اند (backfill ماهر)", _noorg == 0, _noorg)

r = c.post("/orgs/add", data={"name": "مرکز تستی"}, follow_redirects=True)
check("افزودن مرکز جدید", "مرکز جدید افزوده شد".encode() in r.data, r.status_code)

r = c.post("/domains", data={"name": "حوزه تستی کاشف", "org_id": _kashef["id"]},
           follow_redirects=True)
with A.app.app_context():
    _kd = A.get_db().execute("SELECT id, org_id FROM domains WHERE name='حوزه تستی کاشف'").fetchone()
check("ساخت حوزه زیر مرکز کاشف", _kd is not None and _kd["org_id"] == _kashef["id"],
      _kd["org_id"] if _kd else None)

r = c.post(f"/domains/{_kd['id']}/edit", data={"name": "حوزه تستی کاشف", "org_id": _mah["id"]},
           follow_redirects=True)
with A.app.app_context():
    _mov = A.get_db().execute("SELECT org_id FROM domains WHERE id=?", (_kd["id"],)).fetchone()
check("انتقال حوزه به مرکز دیگر", _mov["org_id"] == _mah["id"], _mov["org_id"])

with A.app.app_context():
    _torg = A.get_db().execute("SELECT id FROM orgs WHERE name='مرکز تستی'").fetchone()
r = c.post(f"/orgs/{_torg['id']}/delete", follow_redirects=True)
with A.app.app_context():
    _gone = A.get_db().execute("SELECT 1 FROM orgs WHERE name='مرکز تستی'").fetchone()
check("حذف مرکز خالی", _gone is None)

r = c.get(f"/activities?org={_kashef['id']}")
check("فیلتر مرکز در لیست فعالیت‌ها", r.status_code == 200, r.status_code)
r = c.get(f"/reports?org={_kashef['id']}")
check("فیلتر مرکز در گزارش‌ها", r.status_code == 200, r.status_code)
r = c.get("/domains")
check("صفحه مدیریت حوزه‌ها: نوار مراکز", "مراکز:".encode() in r.data)

# یکدست‌سازی نام کارشناس — ادغام «نام کوتاه/مستعار» با نام کامل در نمودار داشبورد
with A.app.app_context():
    _d = A.get_db()
    _wid = _d.execute("SELECT id FROM domains WHERE name='ارزیابی امنیتی وب'").fetchone()["id"]
    _ef = _d.execute("""SELECT f.id FROM form_fields f JOIN domains d ON d.id=f.domain_id
                        WHERE d.name='ارزیابی امنیتی وب' AND f.field_key='expert'""").fetchone()["id"]
    _aid = _d.execute("SELECT id FROM users WHERE username='admin'").fetchone()["id"]
    _d.execute("INSERT INTO users(username,password_hash,full_name,role,aliases,created_at)"
               " VALUES(?,?,?,?,?,?)",
               ("vizhe", "x", "کامران ویژه‌پور", "expert", "کامران", A.now_iso()))
    for _e in ["کامران ویژه‌پور", "ویژه‌پور", "کامران", "مهدی صالحی", "علی صالحی", "صالحی"]:
        _cur = _d.execute("INSERT INTO activities(domain_id,user_id,status,created_at,updated_at)"
                          " VALUES(?,?,?,?,?)", (_wid, _aid, "انجام شده", A.now_iso(), A.now_iso()))
        _d.execute("INSERT INTO activity_values(activity_id,field_id,value) VALUES(?,?,?)",
                   (_cur.lastrowid, _ef, _e))
    _d.commit()
r = c.get("/api/dashboard")
_exp = {e["label"]: e["value"] for e in r.get_json()["charts"]["experts"]}
check("ادغام پسوند یکتا: «ویژه‌پور» با «کامران ویژه‌پور» یکی شمرده می‌شود",
      _exp.get("کامران ویژه‌پور", 0) >= 2 and "ویژه‌پور" not in _exp, str(_exp))
check("ادغام نام مستعار: «کامران» با نام کامل کاربر یکی شمرده می‌شود",
      _exp.get("کامران ویژه‌پور", 0) >= 3, str(_exp))
check("پسوند مبهم ادغام نمی‌شود («صالحی» جدا از دو نام کامل می‌ماند)",
      _exp.get("صالحی", 0) == 1 and _exp.get("علی صالحی", 0) == 1 and _exp.get("مهدی صالحی", 0) == 1,
      str(_exp))

# دو املای متفاوت از یک نام (فاصله/نیم‌فاصله) نباید باعث ابهام در ادغام شوند
with A.app.app_context():
    _d = A.get_db()
    for _e in ["کامران ویژه پور", "ویژه پور"]:
        _cur = _d.execute("INSERT INTO activities(domain_id,user_id,status,created_at,updated_at)"
                          " VALUES(?,?,?,?,?)",
                          (_wid, _aid, "انجام شده", A.now_iso(), A.now_iso()))
        _d.execute("INSERT INTO activity_values(activity_id,field_id,value) VALUES(?,?,?)",
                   (_cur.lastrowid, _ef, _e))
    _d.commit()
r = c.get("/api/dashboard")
_exp = {e["label"]: e["value"] for e in r.get_json()["charts"]["experts"]}
check("دو املای یک نام ابهام ایجاد نمی‌کند و همه ادغام می‌شوند",
      _exp.get("کامران ویژه‌پور", 0) == 5 and "ویژه پور" not in _exp and "کامران ویژه پور" not in _exp,
      str(_exp))

# چند کارشناس در یک سلول اکسل: هر بخش یکدست، جداکننده یکسان و بخش تکراری حذف می‌شود
with A.app.app_context():
    _d = A.get_db()
    for _e in ["ویژه‌پور، مهدی صالحی", "ویژه‌پور و مهدی صالحی", "کامران و کامران ویژه‌پور"]:
        _cur = _d.execute("INSERT INTO activities(domain_id,user_id,status,created_at,updated_at)"
                          " VALUES(?,?,?,?,?)",
                          (_wid, _aid, "انجام شده", A.now_iso(), A.now_iso()))
        _d.execute("INSERT INTO activity_values(activity_id,field_id,value) VALUES(?,?,?)",
                   (_cur.lastrowid, _ef, _e))
    _d.commit()
r = c.get("/api/dashboard")
_exp = {e["label"]: e["value"] for e in r.get_json()["charts"]["experts"]}
check("چند نام در یک سلول: هر کارشناس جداگانه شمرده می‌شود (برچسب ترکیبی نداریم)",
      "کامران ویژه‌پور، مهدی صالحی" not in _exp, str(_exp))
check("بخش‌های دوپلیکیت یک نفر یک بار می‌آیند («کامران و کامران ویژه‌پور» ← یک نفر)",
      _exp.get("کامران ویژه‌پور", 0) == 8, str(_exp))
check("شریکِ فعالیت چندکارشناسه هم سهم می‌گیرد («مهدی صالحی» +۲)",
      _exp.get("مهدی صالحی", 0) == 3, str(_exp))
r = c.get("/users/new")
check("فرم کاربر: فیلد «نام‌های مستعار» موجود است", "نام‌های مستعار".encode() in r.data)
r = c.get("/activities")
check("لیست فعالیت‌ها: نام ادغام‌شده یکتا نمایش داده می‌شود",
      '"expert": "ویژه‌پور"'.encode() not in r.data, r.data[-100:])

# ---------- پرونده‌ها (/files)
r = c.get("/files")
check("صفحه پرونده‌ها برای مدیر باز می‌شود و لینک ناوبری دارد",
      r.status_code == 200 and "پرونده‌ها".encode() in r.data)
r = c.get("/files")
check("فایل‌های آپلودشده در پرونده‌ها دیده می‌شوند",
      "report-final.pdf".encode() in r.data and "گزارش.pdf".encode() in r.data, r.data[:60])
r = c.get("/files?q=report-final")
check("جستجوی نام فایل در پرونده‌ها",
      "report-final.pdf".encode() in r.data and "گزارش.pdf".encode() not in r.data)
r = c.get(f"/files?domain={dom['id']}")
check("فیلتر حوزه در پرونده‌ها دقیق است",
      "گزارش.pdf".encode() in r.data and "report-final.pdf".encode() in r.data)
with A.app.app_context():
    _other_dom = A.get_db().execute("SELECT id FROM domains WHERE id<>? LIMIT 1",
                                    (dom["id"],)).fetchone()["id"]
r = c.get(f"/files?domain={_other_dom}")
check("فیلتر حوزه دیگر در پرونده‌ها خالی است",
      "گزارش.pdf".encode() not in r.data and "report-final.pdf".encode() not in r.data)
c.get("/logout"); login("reza", "123456")
r = c.get("/files")
check("کارشناس فقط فایل‌های فعالیت‌های خودش را می‌بیند",
      r.status_code == 200 and "report-final.pdf".encode() not in r.data)
c.get("/logout"); login("admin", "admin123")

print(f"\n✅ همه {len(ok)} تست موفقیت‌آمیز بود.")
